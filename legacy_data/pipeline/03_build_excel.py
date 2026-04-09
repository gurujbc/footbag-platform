#!/usr/bin/env python3
"""
03_build_excel.py — DEPRECATED

This script is no longer the workbook deliverable.

It produces a summary-column format (10 rows per event) that does not match
the canonical per-placement year-sheet layout used in the release workbook.

Superseded by:
  pipeline/build_workbook_release.py   — v22-style canonical release workbook (forthcoming)
  pipeline/build_workbook_community.py — community distribution format

Kept for reference / audit traceability. Do not run in production.
--- original docstring below ---

Stage 3: Build final Excel workbook
- Reads out/stage2_canonical_events.csv
- Generates Excel workbook with one sheet per year
- Outputs: Footbag_Results_Canonical.xlsx
"""

from __future__ import annotations

from pathlib import Path
import sys

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import csv
import json
import re
import string
import hashlib
import unicodedata
from copy import copy
from typing import Optional
from collections import defaultdict

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

OUT_DIR = REPO_ROOT / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Import master QC orchestrator
try:
    from qc import qc_master
    USE_MASTER_QC = True
except ImportError:
    print("Warning: Could not import qc.qc_master, Stage 3 QC will not run")
    USE_MASTER_QC = False


# Excel/openpyxl rejects control chars: 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

# Characters that don't decompose cleanly via NFKD; map to closest ASCII equivalent.
_ASCII_PRE_MAP: dict[str, str] = {
    "ł": "l", "Ł": "L", "ø": "o", "Ø": "O", "ß": "ss",
    "đ": "d", "Đ": "D", "ı": "i", "ŋ": "n",
    "þ": "th", "Þ": "Th", "æ": "ae", "Æ": "AE",
    "œ": "oe", "Œ": "OE", "ð": "d", "Ð": "D",
    "\u2013": "-",   # en-dash
    "\u2014": "-",   # em-dash
    "\u2018": "'",   # left single quote
    "\u2019": "'",   # right single quote / apostrophe
    "\u201c": '"',   # left double quote
    "\u201d": '"',   # right double quote
    "\u00b0": "",    # degree sign
    "\ufffd": "?",   # replacement character
}
_ASCII_PRE_TABLE = str.maketrans(_ASCII_PRE_MAP)

# Quarantine event styling (archival workbook)
QUARANTINE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
QUARANTINE_FONT = Font(italic=True, color="767676")


def _to_ascii(s: str) -> str:
    """
    Transliterate a string to plain ASCII for Excel cell output.
    Applies pre-map for non-decomposable characters, then NFKD normalization,
    then drops any remaining non-ASCII bytes.  Newlines and tabs are preserved.
    """
    if not isinstance(s, str):
        return s
    s = s.translate(_ASCII_PRE_TABLE)
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if ord(c) < 128 or c in "\n\t")


def load_alias_map(path):
    """
    Read-only alias map.
    Returns: dict[player_id] -> alias_group_id

    No guessing.
    No transitive closure.
    Blank means 'unknown / not yet reviewed'.
    """
    import csv
    from pathlib import Path

    alias_map = {}
    p = Path(path)

    if not p.exists():
        return alias_map

    with p.open("r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            pid = row.get("player_id", "").strip()
            agid = row.get("alias_group_id", "").strip()

            if pid and agid:
                alias_map[pid] = agid

    return alias_map


ALIAS_MAP = load_alias_map(OUT_DIR / "person_alias_map_bootstrap.csv")


def sanitize_excel_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Sanitize all string cells for Excel: strip control chars + transliterate to ASCII."""
    if df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_string_dtype(out[col]) or out[col].dtype == object:
            out[col] = out[col].apply(
                lambda v: _ILLEGAL_XLSX_RE.sub("", _to_ascii(v)) if isinstance(v, str) else v
            )
    return out


def sanitize_string(s: str) -> str:
    """Sanitize a single string for Excel: strip control chars + transliterate to ASCII."""
    if not isinstance(s, str):
        return s
    return _ILLEGAL_XLSX_RE.sub("", _to_ascii(s))


def _strip_diacritics(s: str) -> str:
    if not isinstance(s, str):
        return ""
    # NFKD splits accents; we drop combining marks
    return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))


def normalize_person_key(name: str) -> str:
    """
    Conservative, presentation-only normalization key for alias-candidate grouping.
    Lowercase, strip diacritics, remove punctuation, collapse whitespace.
    """
    if not isinstance(name, str) or not name.strip():
        return ""
    t = _strip_diacritics(name).lower()
    t = re.sub(r"[^a-z0-9\s]", " ", t)   # remove punctuation/symbols
    t = re.sub(r"\s+", " ", t).strip()
    # Optional: remove single-letter middle initials (keeps first + last)
    t = re.sub(r"\b([a-z])\b", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _stable_group_id(prefix: str, key: str) -> str:
    h = hashlib.sha1(f"{prefix}:{key}".encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{h}"


def _best_display_name(names: list[str]) -> str:
    cleaned = [n.strip() for n in names if isinstance(n, str) and n.strip()]
    if not cleaned:
        return ""
    # Prefer shortest (usually most canonical-looking), then alpha
    cleaned.sort(key=lambda x: (len(x), x.lower()))
    return cleaned[0]


def _alias_confidence(alias_names: list[str], key: str) -> str:
    """
    Simple heuristic:
    - high: all aliases normalize to same key
    - med: >1 alias and share same last token after normalization
    - low: otherwise
    """
    normed = [normalize_person_key(a) for a in alias_names if isinstance(a, str)]
    normed = [n for n in normed if n]
    if normed and all(n == key for n in normed):
        return "high"
    # last-token check
    toks = [n.split() for n in normed if n.split()]
    lasts = [t[-1] for t in toks if t]
    if len(set(lasts)) == 1 and len(lasts) >= 2:
        return "med"
    return "low"


def normalize_team_key(p1: str, p2: str) -> str:
    """
    Presentation-only team grouping key based on *member names* (not IDs),
    order-invariant: (A,B) == (B,A).
    """
    a = normalize_person_key(p1)
    b = normalize_person_key(p2)
    if not a or not b:
        return ""
    left, right = sorted([a, b])
    return f"{left} // {right}"


def _best_team_display(alias_pairs: list[str]) -> str:
    """
    Choose a stable best display label for team: shortest alias pair string.
    """
    cleaned = [s.strip() for s in alias_pairs if isinstance(s, str) and s.strip()]
    if not cleaned:
        return ""
    cleaned.sort(key=lambda x: (len(x), x.lower()))
    return cleaned[0]


def team_display_name(n1: str, n2: str) -> str:
    a = (n1 or "").strip()
    b = (n2 or "").strip()
    if not a or not b:
        return a or b
    return " / ".join(sorted([a, b], key=lambda x: x.lower()))


def _collapse_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def normalize_division_key(s: str) -> str:
    """
    Presentation-only key to group near-identical divisions for the Divisions_Normalized sheet.
    Conservative: remove punctuation + collapse whitespace + lowercase.
    Also normalizes trailing 'competition/comp' to reduce obvious redundancy
    (e.g., 'Freestyle Circle' vs 'Freestyle Circle Competition').
    """
    if not isinstance(s, str) or not s.strip():
        return ""
    t = s.lower().strip()
    # Replace punctuation with spaces (keeps words separated)
    trans = str.maketrans({ch: " " for ch in string.punctuation})
    t = t.translate(trans)
    t = _collapse_ws(t)
    # Common harmless suffix noise
    t = re.sub(r"\bcompetition\b$", "", t).strip()
    t = re.sub(r"\bcomp\b$", "", t).strip()
    t = _collapse_ws(t)
    return t


def pick_best_division_display(aliases: list[str]) -> str:
    """
    Choose a human-friendly label from observed aliases.
    Heuristic: prefer shorter, non-empty; preserve original capitalization.
    """
    cleaned = [a.strip() for a in (aliases or []) if isinstance(a, str) and a.strip()]
    if not cleaned:
        return ""
    # Prefer the shortest variant (often the canonical-looking one)
    cleaned.sort(key=lambda x: (len(x), x.lower()))
    return cleaned[0]


def year_to_sheet_name(y) -> str:
    if y is None:
        return "Unknown"
    try:
        # Handles "2001.0", 2001.0, "2001"
        yi = int(float(str(y).strip()))
        return str(yi)
    except Exception:
        return str(y)


def _parse_date_for_sort(date_str: str) -> tuple[int, int, int]:
    """
    Parse date string to (year, month, day) for chronological sorting.
    Returns (9999, 12, 31) if unparseable so such events sort last within the year.
    """
    if not date_str or not str(date_str).strip():
        return (9999, 12, 31)
    s = str(date_str).strip()
    # ISO-style YYYY-MM-DD
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # Year only (19xx or 20xx)
    year_m = re.search(r"\b(19|20)(\d{2})\b", s)
    if year_m:
        y = int(year_m.group(1) + year_m.group(2))
        # Try "Month DD" or "Month D" before year
        month_names = r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*"
        md_m = re.search(rf"({month_names})\s+(\d{{1,2}})", s, re.IGNORECASE)
        if md_m:
            month_map = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                         "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
            mon_str = md_m.group(1)[:3].lower()
            month = month_map.get(mon_str, 0)
            day = int(md_m.group(2))
            return (y, month, day)
        return (y, 0, 0)
    return (9999, 12, 31)


def display_date(date_str: str, year) -> str:
    """
    Presentation-only: if date has no explicit YYYY and we *know* the sheet year,
    append ', YYYY' to improve consistency across the final workbook.
    """
    s = sanitize_string(date_str or "")
    if not s:
        return ""
    if year is None:
        return s
    # If a 4-digit year is already present, leave it alone
    if re.search(r"\b(19|20)\d{2}\b", s):
        return s
    try:
        y = int(year)
    except Exception:
        return s
    return f"{s}, {y}"


def is_team_division(division_name: str) -> bool:
    """
    Determine if a division is a team division based on division name.
    
    Returns True if division contains team indicators (doubles, pairs, team),
    False if it contains "singles" or is empty/None.
    """
    if not division_name:
        return False

    name = division_name.lower()

    # Explicit singles exclusions
    if "singles" in name:
        return False

    # Explicit team indicators
    if any(k in name for k in ["doubles", "pairs", "team"]):
        return True

    return False


# ------------------------------------------------------------
# Results formatting from placements
# ------------------------------------------------------------

# Category display order and labels
CATEGORY_ORDER = ["net", "freestyle", "golf", "sideline", "unknown"]
CATEGORY_LABELS = {
    "net": "NET",
    "freestyle": "FREESTYLE",
    "golf": "GOLF",
    "sideline": "OTHER",
    "unknown": "OTHER",
}


def format_results_from_placements(placements: list[dict], players_by_id: Optional[dict] = None) -> Optional[str]:
    """
    Build a deterministic, consistent results blob from canonical placements.
    Groups results by category (NET, FREESTYLE, GOLF, OTHER) with clear headers.

    Format:
      === NET ===
      OPEN SINGLES NET
      1. Name
      2. Name / Name

      === FREESTYLE ===
      SHRED 30
      1. Name

    We do NOT invent missing facts. If no placements exist -> None.
    """
    if not placements:
        return None

    # Group by category, then by division
    by_category = {}
    for p in placements:
        cat = p.get("division_category", "unknown") or "unknown"
        div = p.get("division_canon") or p.get("division_raw") or "Unknown"

        if cat not in by_category:
            by_category[cat] = {}
        if div not in by_category[cat]:
            by_category[cat][div] = []
        by_category[cat][div].append(p)

    out_lines = []

    # Output categories in defined order
    for cat in CATEGORY_ORDER:
        if cat not in by_category:
            continue

        divisions = by_category[cat]
        if not divisions:
            continue

        # Add category header
        label = CATEGORY_LABELS.get(cat, cat.upper())
        out_lines.append(f"<<< {label} >>>")
        out_lines.append("")

        # Sort divisions alphabetically within category
        for div in sorted(divisions.keys(), key=str.casefold):
            entries = divisions[div]

            # Sort entries by place, then by player name
            def sort_key(p):
                place = p.get("place", 999)
                try:
                    place = int(place)
                except (ValueError, TypeError):
                    place = 999
                name = _build_name_line(p, players_by_id)
                return (place, name.lower() if name else "")

            entries.sort(key=sort_key)

            s_ref = (entries[0].get("source_ref", "") or "").strip() if entries else ""
            header = f"--- {div.upper()} ---"
            if s_ref:
                header += f" (Source: {s_ref})"
            out_lines.append(header)

            # Deduplicate: skip (place, name) combos already output (source data can have dupes)
            seen_line_key = set()
            for p in entries:
                place = p.get("place")
                try:
                    place_int = int(place)
                    place_txt = f"{place_int}."
                except (ValueError, TypeError):
                    place_txt = f"{place}." if place is not None else ""

                name = _build_name_line(p, players_by_id)
                line_key = (place_txt, (name or "").lower().strip())
                if line_key in seen_line_key:
                    continue
                seen_line_key.add(line_key)

                if place_txt:
                    out_lines.append(f"{place_txt} {name}".rstrip())
                else:
                    out_lines.append(name)

            out_lines.append("")  # blank line between divisions

    # Remove trailing blank lines
    while out_lines and out_lines[-1] == "":
        out_lines.pop()

    return "\n".join(out_lines) if out_lines else None


def _build_name_line(placement: dict, players_by_id: Optional[dict] = None) -> str:
    """Build display name using modern column names (person_id, person_canon)."""
    def _lookup_clean(which: str) -> str:
        if which == "player1":
            # Check every possible ID/Name column name used in the various pipeline stages
            pid = placement.get("person_id") or placement.get("player_id") or placement.get("player1_id") or ""
            raw = placement.get("person_canon") or placement.get("player_name_clean") or placement.get("player1_name") or ""
        else:
            pid = placement.get("player2_id") or ""
            raw = placement.get("player2_name") or ""

        # If we have a canonical mapping from our truth file, use it
        if players_by_id and pid and pid in players_by_id:
            return (players_by_id[pid].get('player_name_clean') or raw).strip()
        return str(raw).strip()

    p1 = _lookup_clean("player1")
    p2 = _lookup_clean("player2")
    return f"{p1} / {p2}".strip(" /") if p2 else p1


# ------------------------------------------------------------
# CSV reading
# ------------------------------------------------------------
def _read_optional_csv(path: Path) -> pd.DataFrame:
    if path.exists():
        return pd.read_csv(path, dtype=str).fillna("")
    return pd.DataFrame()


def read_stage2_csv(csv_path: Path) -> list[dict]:
    """Read stage2 CSV and return list of event records."""
    # Increase CSV field size limit to handle large JSON fields
    csv.field_size_limit(min(2**31 - 1, 10 * 1024 * 1024))  # 10MB limit
    _NAN_STRINGS = {"nan", "none", "null", "na", "#n/a"}
    records = []
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Normalise pandas/Excel NaN sentinel strings to empty string
            for key, val in row.items():
                if isinstance(val, str) and val.strip().lower() in _NAN_STRINGS:
                    row[key] = ""
            # Convert year to int if present
            if row.get("year"):
                try:
                    row["year"] = int(row["year"])
                except ValueError:
                    row["year"] = None
            else:
                row["year"] = None

            # Parse placements JSON
            placements_json = row.get("placements_json", "[]")
            try:
                row["placements"] = json.loads(placements_json)
            except json.JSONDecodeError:
                row["placements"] = []

            records.append(row)
    return records


def build_players_by_id(players_df: Optional[pd.DataFrame]) -> dict:
    """Build lookup: player_id -> {player_name_clean, country_clean, name_status}."""
    if players_df is None or len(players_df) == 0:
        return {}
    dfp = players_df.copy()
    # Filter out junk rows so they cannot leak into Excel outputs
    if 'name_status' in dfp.columns:
        dfp = dfp[dfp['name_status'].isin(['ok','suspicious','needs_review'])].copy()
    out = {}
    for _, r in dfp.iterrows():
        pid = str(r.get('player_id') or '').strip()
        if not pid:
            continue
        out[pid] = {
            'player_name_clean': str(r.get('player_name_clean') or r.get('player_name_raw') or '').strip(),
            'country_clean': str(r.get('country_clean') or r.get('country_observed') or '').strip(),
            'name_status': str(r.get('name_status') or '').strip(),
        }
    return out


_WS = re.compile(r"\s+")


def _one_line(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    # collapse any \r \n \t etc into single spaces
    return _WS.sub(" ", s).strip()


def build_persons_truth(df_pf: pd.DataFrame) -> pd.DataFrame:
    """
    Build Persons_Truth table:
    - one row per effective_person_id
    - NO guessing
    - person_id from overrides OR fallback to player_id
    """
    if df_pf is None or (isinstance(df_pf, pd.DataFrame) and df_pf.empty):
        return pd.DataFrame()

    rows = []

    for _, r in df_pf.iterrows():
        pid = str(r.get("person_id") or "").strip()
        player_id = str(r.get("player_id") or r.get("player1_id") or "").strip()

        if pid:
            effective_person_id = pid
            source = "override"
        elif player_id:
            effective_person_id = player_id
            source = "fallback_player_id"
        else:
            continue  # impossible case

        rows.append({
            "effective_person_id": effective_person_id,
            "source": source,
            "player_id": player_id,
            "player_name_clean": str(r.get("player_name_clean") or "").strip(),
        })

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    # Aggregate deterministically
    out = []
    for ep, g in df.groupby("effective_person_id"):
        names = sorted({n for n in g["player_name_clean"] if n})
        player_ids = sorted({p for p in g["player_id"] if p})

        out.append({
            "person_id": ep,
            "identity_source": g["source"].iloc[0],
            "player_ids_seen": " | ".join(player_ids),
            "player_names_seen": " | ".join(names),
            "player_id_count": len(player_ids),
            "name_variant_count": len(names),
        })

    return pd.DataFrame(out).sort_values(
        by=["identity_source", "name_variant_count"],
        ascending=[True, False],
    )


# ------------------------------------------------------------
# Excel writer
# ------------------------------------------------------------
def write_excel(
    out_xlsx: Path,
    records: list[dict],
    players_df: Optional[pd.DataFrame] = None,
    placements_flat_df: Optional[pd.DataFrame] = None,
    unresolved_df: Optional[pd.DataFrame] = None,
    events_df: Optional[pd.DataFrame] = None,
    results_map: Optional[dict] = None,
) -> None:
    """
    Archive workbook writer (matches Footbag_Results_Canonical.xlsx layout):
    - One sheet per year named YYYY.0
    - Columns are event_id
    - Rows are fixed labels (Tournament Name, Date, Location, ...)
    - Results are generated from placements (canonical), not copied raw
    - If events_df is provided, sheets are grouped by year from events_df and
      per-event placement counts come from placements_flat_df filtered by event_id.
    """
    players_by_id = build_players_by_id(players_df)

    def _loc(eid: str, rec: dict) -> str:
        """Return location display string from event record."""
        return rec.get("location") or ""

    # Use pre-built results_map when provided (e.g. from flat placements); else build from records
    if results_map is None:
        results_map = {}
        for rec in records:
            eid = rec.get("event_id")
            if eid:
                placements = rec.get("placements", [])
                results_map[str(eid)] = format_results_from_placements(placements, players_by_id)

    # Fixed row labels (index) to match the example workbook
    row_labels = [
        "Tournament Name",
        "Date",
        "Location",
        "Event Type",
        "Host Club",
        "Results",
        "Source Ref",
        "Ver",
    ]

    # Sort key for event IDs
    def _eid_sort_key(x: str):
        try:
            return int(re.sub(r"\D+", "", x) or "0")
        except Exception:
            return 0

    def _chronological_sort_key(rec: dict) -> tuple:
        """Sort by date (year, month, day), then by event_id."""
        date_tuple = _parse_date_for_sort(rec.get("date") or "")
        eid = str(rec.get("event_id", ""))
        return (date_tuple[0], date_tuple[1], date_tuple[2], _eid_sort_key(eid))

    # Group by year: from events_df when provided, otherwise from records
    by_year = {}
    unknown_year = []
    records_by_eid = {str(rec.get("event_id", "")): rec for rec in records if rec.get("event_id")}

    if events_df is not None and not events_df.empty and "year" in events_df.columns:
        for year, group in events_df.groupby("year"):
            try:
                year_val = int(float(year)) if pd.notna(year) and str(year).strip() else None
            except (ValueError, TypeError):
                year_val = None
            year_records_list = []
            for _, event in group.iterrows():
                eid = str(event.get("event_id", ""))
                if not eid:
                    continue
                rec = records_by_eid.get(eid)
                if rec is not None:
                    year_records_list.append(rec)
            if year_val is not None:
                by_year[year_val] = year_records_list
            else:
                unknown_year.extend(year_records_list)
    else:
        for rec in records:
            year = rec.get("year")
            if year is not None:
                if year not in by_year:
                    by_year[year] = []
                by_year[year].append(rec)
            else:
                unknown_year.append(rec)

    event_locator = {}  # event_id(str) -> (sheet_name(str), col_idx(int))

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        # Build one sheet per year (events ordered chronologically by date)
        for y in sorted(by_year.keys()):
            year_records = sorted(by_year[y], key=_chronological_sort_key)
            eids = [str(r.get("event_id", "")) for r in year_records]

            # Excel columns: A=1 is row labels, so first event column is B=2
            sheet_name = year_to_sheet_name(y)
            for j, eid in enumerate(eids, start=2):
                event_locator[str(eid)] = (sheet_name, j)

            data = {}
            for eid in eids:
                rec = next((r for r in year_records if str(r.get("event_id")) == eid), None)
                if not rec:
                    continue

                # Use integer event_id as column header to avoid Excel apostrophe prefix
                col_key = int(eid) if eid.isdigit() else eid
                source_ref = rec.get("source_ref", "") or ""
                ver_level = str(rec.get("verification_level", "2"))
                data[col_key] = [
                    sanitize_string(rec.get("event_name") or ""),
                    display_date(rec.get("date") if pd.notna(rec.get("date")) else "", y),
                    sanitize_string(_loc(eid, rec)),
                    sanitize_string(rec.get("event_type") or ""),
                    sanitize_string(rec.get("host_club") or ""),
                    sanitize_string(results_map.get(eid) or ""),
                    sanitize_string(source_ref),
                    ver_level,
                ]

            df_year = pd.DataFrame(data, index=row_labels)
            df_year.index.name = "event_id"  # puts "event_id" in A1 like the example

            df_year = sanitize_excel_strings(df_year)
            df_year.to_excel(xw, sheet_name=sheet_name)

            worksheet = xw.sheets[sheet_name]

            # Insert a year-banner row above the event-ID header row.
            # After insert: row 1 = banner, row 2 = event IDs, rows 3-8 = data.
            worksheet.insert_rows(1)
            n_cols = len(eids) + 1  # col A (labels) + one per event
            last_col = get_column_letter(n_cols)
            worksheet.merge_cells(f"A1:{last_col}1")
            banner = worksheet["A1"]
            banner.value = int(y)
            banner.alignment = Alignment(horizontal="center", vertical="center")

            # Apply wrap_text formatting to Results row (now row 8 after banner insert)
            for col_idx in range(2, len(eids) + 2):  # Start from column B (2)
                cell = worksheet.cell(row=8, column=col_idx)
                a = copy(cell.alignment)
                a.wrap_text = True
                cell.alignment = a
            # Apply quarantine styling to event columns where status is 'quarantine'
            for col_idx, eid in enumerate(eids, start=2):
                rec = next((r for r in year_records if str(r.get("event_id")) == eid), None)
                if rec and rec.get("status") == "quarantine":
                    for row_idx in range(3, 3 + len(row_labels)):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = QUARANTINE_FILL
                        cell.font = QUARANTINE_FONT
            # Freeze: banner + event-ID header row visible; col A (labels) visible
            worksheet.freeze_panes = "B3"

        # Unknown-year sheet (chronological by date, then event_id)
        if unknown_year:
            unknown_sorted = sorted(unknown_year, key=_chronological_sort_key)
            eids = [str(r.get("event_id", "")) for r in unknown_sorted]
            # Excel columns: A=1 is row labels, so first event column is B=2
            for j, eid in enumerate(eids, start=2):
                event_locator[str(eid)] = ("unknown_year", j)
            data = {}
            for eid in eids:
                rec = next((r for r in unknown_year if str(r.get("event_id")) == eid), None)
                if not rec:
                    continue

                # Use integer event_id as column header to avoid Excel apostrophe prefix
                col_key = int(eid) if eid.isdigit() else eid
                source_ref = rec.get("source_ref", "") or ""
                ver_level = str(rec.get("verification_level", "2"))
                data[col_key] = [
                    sanitize_string(rec.get("event_name") or ""),
                    display_date(rec.get("date") or "", None),
                    sanitize_string(_loc(eid, rec)),
                    sanitize_string(rec.get("event_type") or ""),
                    sanitize_string(rec.get("host_club") or ""),
                    sanitize_string(results_map.get(eid) or ""),
                    sanitize_string(source_ref),
                    ver_level,
                ]

            df_unk = pd.DataFrame(data, index=row_labels)
            df_unk.index.name = "event_id"
            df_unk = sanitize_excel_strings(df_unk)
            df_unk.to_excel(xw, sheet_name="unknown_year")

            # Apply wrap_text formatting to Results row
            worksheet = xw.sheets["unknown_year"]
            for col_idx in range(2, len(eids) + 2):
                cell = worksheet.cell(row=7, column=col_idx)
                a = copy(cell.alignment)
                a.wrap_text = True
                cell.alignment = a
            # Apply quarantine styling to event columns where status is 'quarantine'
            for col_idx, eid in enumerate(eids, start=2):
                rec = next((r for r in unknown_year if str(r.get("event_id")) == eid), None)
                if rec and rec.get("status") == "quarantine":
                    for row_idx in range(2, 2 + len(row_labels)):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = QUARANTINE_FILL
                        cell.font = QUARANTINE_FONT

        # Build Index sheet: one row per event with hyperlinks
        index_data = []
        sce_event_ids = {str(rec.get("event_id", "")) for rec in records if rec.get("event_id")}
        for rec in records:
            eid = str(rec.get("event_id", ""))
            if not eid:
                continue

            year = rec.get("year")
            placements = rec.get("placements", [])
            results_text = results_map.get(eid, "")
            results_lines = len(results_text.splitlines()) if results_text else 0

            index_data.append({
                "event_id": eid,
                "year": year if year is not None else "",
                "Tournament Name": sanitize_string(rec.get("event_name") or ""),
                "Date": sanitize_string(rec.get("date") or ""),
                "Location": sanitize_string(_loc(eid, rec)),
                "Event Type": sanitize_string(rec.get("event_type") or ""),
                "Host Club": sanitize_string(rec.get("host_club") or ""),
                "placements_count": len(placements),
                "results_lines": results_lines,
                "event_source": "",
            })

        # Build PBP row-count lookup: used to override placements_count so the Index
        # reflects what is actually queryable in Placements_ByPerson, not the stage2 raw
        # count (which can include noise/unresolvable entries that never reach PBP).
        pf_count_by_event: dict[str, int] = {}
        if placements_flat_df is not None and not placements_flat_df.empty:
            for _eid, _grp in placements_flat_df.groupby(placements_flat_df["event_id"].astype(str)):
                pf_count_by_event[str(_eid)] = len(_grp)

        # Back-fill placements_count in the already-built index_data rows using PBP counts.
        for row in index_data:
            row["placements_count"] = pf_count_by_event.get(str(row["event_id"]), 0)

        # Append Index-only stub rows for synthetic events in PBP but not in stage2.
        # These are pre-mirror historical events (1980-1986) and a few others whose
        # results were compiled from non-mirror sources.
        if placements_flat_df is not None and not placements_flat_df.empty:
            pf_event_ids = placements_flat_df["event_id"].dropna().astype(str).unique()
            for eid in sorted(pf_event_ids, key=_eid_sort_key):
                if eid in sce_event_ids:
                    continue
                pf_rows = placements_flat_df[placements_flat_df["event_id"].astype(str) == eid]
                year_val = int(pf_rows["year"].iloc[0]) if len(pf_rows) > 0 else None
                cats = sorted(pf_rows["division_category"].dropna().unique()) if "division_category" in pf_rows.columns else []
                index_data.append({
                    "event_id": eid,
                    "year": year_val if year_val is not None else "",
                    "Tournament Name": f"(pre-mirror event, {year_val})" if year_val else "(pre-mirror event)",
                    "Date": "",
                    "Location": "",
                    "Event Type": ",".join(cats),
                    "Host Club": "",
                    "placements_count": len(pf_rows),
                    "results_lines": 0,
                    "event_source": "synthetic_pre_mirror",
                })

        # Sort by year, then event_id
        index_data.sort(key=lambda x: (x["year"] if x["year"] != "" else 9999, _eid_sort_key(x["event_id"])))
        
        df_index = pd.DataFrame(index_data)
        df_index = sanitize_excel_strings(df_index)
        df_index.to_excel(xw, sheet_name="Index", index=False)
        
        # Apply formatting to Index sheet: hyperlinks, filters, freeze panes
        index_ws = xw.sheets["Index"]
        
        # Add hyperlinks to event_id column (column A, starting at row 2)
        hyperlink_font = Font(color="0563C1", underline="single")  # Blue, underlined
        for idx, row_data in enumerate(index_data, start=2):
            eid = row_data["event_id"]
            if eid in event_locator:
                sheet_name, col_idx = event_locator[eid]
                col_letter = get_column_letter(col_idx)
                # Hyperlink format: #SheetName!ColumnLetter1 (e.g., #1999!B1)
                hyperlink = f"#{sheet_name}!{col_letter}1"
                cell = index_ws.cell(row=idx, column=1)  # Column A
                cell.hyperlink = hyperlink
                cell.font = hyperlink_font
        
        # Freeze first row (header)
        index_ws.freeze_panes = "A2"
        
        # Add auto filter to header row
        index_ws.auto_filter.ref = index_ws.dimensions

        # Build Summary sheet: rollups and health metrics
        total_events = len(records)
        if events_df is not None and "year" in events_df.columns and placements_flat_df is not None and not placements_flat_df.empty:
            total_placements = len(placements_flat_df)
            years_with_events = []
            year_stats = defaultdict(lambda: {"events": 0, "placements": 0})
            for year, group in events_df.groupby("year"):
                try:
                    year_val = int(float(year)) if pd.notna(year) and str(year).strip() else None
                except (ValueError, TypeError):
                    year_val = None
                if year_val is not None:
                    years_with_events.append(year_val)
                for _, event in group.iterrows():
                    eid = str(event.get("event_id", ""))
                    if not eid:
                        continue
                    if year_val is not None:
                        year_stats[year_val]["events"] += 1
                    this_event_placements = placements_flat_df[placements_flat_df["event_id"] == eid]
                    placement_count = len(this_event_placements)
                    if year_val is not None:
                        year_stats[year_val]["placements"] += placement_count
        else:
            total_placements = sum(len(rec.get("placements", [])) for rec in records)
            years_with_events = [rec.get("year") for rec in records if rec.get("year") is not None]
            year_stats = defaultdict(lambda: {"events": 0, "placements": 0})
            for rec in records:
                year = rec.get("year")
                if year is not None:
                    year_stats[year]["events"] += 1
                    year_stats[year]["placements"] += len(rec.get("placements", []))

        year_min = min(years_with_events) if years_with_events else None
        year_max = max(years_with_events) if years_with_events else None
        
        pf_rows = len(placements_flat_df) if placements_flat_df is not None else 0
        pf_resolved = (
            placements_flat_df["person_unresolved"].fillna("").astype(str).str.strip().str.lower().ne("true").sum()
            if placements_flat_df is not None and "person_unresolved" in placements_flat_df.columns
            else pf_rows
        )

        summary_data = [
            {"Metric": "Total Events", "Value": total_events},
            {"Metric": "Total Placements (stage2 raw)", "Value": total_placements},
            {"Metric": "Total Placements (identity-locked / PBP)", "Value": pf_rows},
            {"Metric": "Total Placements (resolved persons only)", "Value": pf_resolved},
            {"Metric": "Year Min", "Value": year_min if year_min else ""},
            {"Metric": "Year Max", "Value": year_max if year_max else ""},
        ]
        
        df_summary_metrics = pd.DataFrame(summary_data)
        df_summary_metrics = sanitize_excel_strings(df_summary_metrics)
        
        year_table_data = [
            {"year": year, "events": stats["events"], "placements": stats["placements"]}
            for year, stats in sorted(year_stats.items())
        ]
        df_year_table = pd.DataFrame(year_table_data)
        df_year_table = sanitize_excel_strings(df_year_table)
        
        # Write Summary sheet with metrics and year table
        df_summary_metrics.to_excel(xw, sheet_name="Summary", index=False, startrow=0)
        df_year_table.to_excel(xw, sheet_name="Summary", index=False, startrow=len(summary_data) + 3)
        
        summary_ws = xw.sheets["Summary"]
        summary_ws.freeze_panes = None
        summary_ws.auto_filter.ref = None

        # ------------------------------------------------------------
        # Players sheets (prefer Stage 2.5 cleaned player tokens)
        # ------------------------------------------------------------
        if players_df is not None and len(players_df) > 0:
            dfp = players_df.copy()

            # Ensure required columns exist (defensive)
            for col in [
                'player_id','player_name_raw','country_observed','player_name_clean',
                'name_status','junk_reason','country_clean','usage_count','source_hint','name_key'
            ]:
                if col not in dfp.columns:
                    dfp[col] = ''

            # Main Players sheet: one row per player_id (truth-preserving, duplicates allowed)
            df_players = dfp[dfp['name_status'].isin(['ok','suspicious','needs_review'])].copy()
            df_players = df_players[['player_id','player_name_clean','country_clean','name_status']]
            df_players = sanitize_excel_strings(df_players)
            df_players.to_excel(xw, sheet_name='Players', index=False)

            # Players_Clean: ok + suspicious (plus audit columns)
            df_clean = dfp[dfp['name_status'].isin(['ok','suspicious'])].copy()
            df_clean = df_clean[['player_id','player_name_clean','country_clean','name_status',
                                 'player_name_raw','country_observed','usage_count','source_hint','name_key']]
            df_clean = sanitize_excel_strings(df_clean)
            df_clean.to_excel(xw, sheet_name='Players_Clean', index=False)

            # Players_Junk: junk only (audit)
            df_junk = dfp[dfp['name_status'] == 'junk'].copy()
            df_junk = df_junk[['player_id','player_name_raw','junk_reason','usage_count','source_hint','name_key']]
            df_junk = sanitize_excel_strings(df_junk)
            df_junk.to_excel(xw, sheet_name='Players_Junk', index=False)

            # ------------------------------------------------------------
            # Players_Alias_Candidates (presentation-only; NO merges)
            # Groups name variants that currently map to different player_id.
            # ------------------------------------------------------------
            df_nonjunk = dfp[dfp['name_status'].isin(['ok','suspicious','needs_review'])].copy()

            # Build grouping key:
            # - prefer provided name_key (Stage 2.5)
            # - fallback to normalized player_name_clean or raw
            def _row_group_key(r):
                nk = r.get('name_key')
                if isinstance(nk, str) and nk.strip():
                    return nk.strip()
                nm = r.get('player_name_clean')
                if not (isinstance(nm, str) and nm.strip()):
                    nm = r.get('player_name_raw')
                return normalize_person_key(nm)

            df_nonjunk['alias_group_key'] = df_nonjunk.apply(_row_group_key, axis=1)
            df_nonjunk['alias_group_key'] = df_nonjunk['alias_group_key'].fillna('').astype(str)

            # Only keep meaningful keys
            df_nonjunk = df_nonjunk[df_nonjunk['alias_group_key'].str.len() > 0]

            candidates_rows = []
            for gkey, g in df_nonjunk.groupby('alias_group_key'):
                # Only interesting if multiple distinct player_ids (your stated problem)
                player_ids = sorted(set([str(x) for x in g['player_id'].dropna().astype(str).tolist() if str(x).strip()]))
                if len(player_ids) < 2:
                    continue

                clean_names = g['player_name_clean'].dropna().astype(str).tolist()
                raw_names = g['player_name_raw'].dropna().astype(str).tolist()
                # Prefer clean names for display; keep raw names as additional aliases if different
                aliases = []
                for n in clean_names:
                    if isinstance(n, str) and n.strip():
                        aliases.append(n.strip())
                for n in raw_names:
                    if isinstance(n, str) and n.strip():
                        aliases.append(n.strip())
                # unique while preserving stable order
                seen = set()
                aliases_u = []
                for a in aliases:
                    if a.lower() not in seen:
                        seen.add(a.lower())
                        aliases_u.append(a)

                display_best = _best_display_name([a for a in aliases_u if a])
                conf = _alias_confidence(aliases_u, normalize_person_key(display_best) or gkey)

                usage = 0
                if 'usage_count' in g.columns:
                    try:
                        usage = int(pd.to_numeric(g['usage_count'], errors='coerce').fillna(0).sum())
                    except Exception:
                        usage = 0

                # country rollup (optional)
                countries = []
                if 'country_clean' in g.columns:
                    countries = sorted(set([str(x).strip() for x in g['country_clean'].dropna().astype(str).tolist() if str(x).strip()]))

                candidates_rows.append({
                    "candidate_group_id": _stable_group_id("p", gkey),
                    "name_key": gkey,
                    "display_name_best": display_best,
                    "aliases": " | ".join(aliases_u),
                    "player_ids": " | ".join(player_ids),
                    "countries_seen": " | ".join(countries),
                    "usage_count_total": usage,
                    "confidence": conf,
                    "review_priority": (
                        "high" if conf == "high" and usage >= 10
                        else "med" if conf in ("high", "med")
                        else "low"
                    ),
                    "merge_decision": "",   # merge / not_merge / unsure
                    "notes": "",
                })

            if candidates_rows:
                # High usage first, then confidence, then name
                conf_rank = {"high": 0, "med": 1, "low": 2}
                candidates_rows.sort(key=lambda r: (-int(r.get("usage_count_total") or 0),
                                                   conf_rank.get(r.get("confidence"), 9),
                                                   (r.get("display_name_best") or "").lower()))
                df_cand = pd.DataFrame(candidates_rows)
                df_cand = sanitize_excel_strings(df_cand)
                df_cand.to_excel(xw, sheet_name="Players_Alias_Candidates", index=False)

                ws = xw.sheets["Players_Alias_Candidates"]
                for col_name in ("player_ids", "name_key"):
                    if col_name in df_cand.columns:
                        idx = df_cand.columns.get_loc(col_name) + 1
                        ws.column_dimensions[get_column_letter(idx)].hidden = True
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        else:
            # Fallback (legacy): derive Players from placements (may include slop)
            # Deterministic improvement: if Placements_Flat has player_name_clean, prefer it for display.
            clean_name_by_player_id: dict[str, str] = {}
            if placements_flat_df is not None and isinstance(placements_flat_df, pd.DataFrame) and not placements_flat_df.empty:
                # Support common id/name column spellings; do NOT invent columns if missing.
                id_col = None
                name_col = None
                for c in ("player_id", "player1_id"):
                    if c in placements_flat_df.columns:
                        id_col = c
                        break
                for c in ("player_name_clean", "player1_name_clean"):
                    if c in placements_flat_df.columns:
                        name_col = c
                        break
                if id_col and name_col:
                    for _pid, _nm in zip(placements_flat_df[id_col].astype(str), placements_flat_df[name_col].astype(str)):
                        _pid = (_pid or "").strip()
                        _nm = (_nm or "").strip()
                        if _pid and _nm and _pid not in clean_name_by_player_id:
                            clean_name_by_player_id[_pid] = _nm

            players_map = {}  # player_id or name -> {name, country}
            for rec in records:
                placements = rec.get('placements', [])
                for p in placements:
                    player_id = (p.get('player1_id') or p.get('player_id') or p.get('player1_player_id') or '')
                    player_name = (p.get('player1_name') or '').strip()
                    if not player_name:
                        continue
                    country = (p.get('home_country') or p.get('country') or p.get('nation') or p.get('player1_country') or p.get('player1_home_country') or '')
                    country = country.strip() if country else ''
                    key = player_id if player_id else player_name.lower()
                    if key not in players_map:
                        players_map[key] = {'player_id': player_id, 'player_name': player_name, 'home_country': country}

            # Prefer clean names for display when available
            for key, val in players_map.items():
                pid = val.get("player_id") or ""
                if pid and pid in clean_name_by_player_id:
                    val["player_name"] = clean_name_by_player_id[pid]

            players_data = list(players_map.values())
            players_data.sort(key=lambda x: (x.get('player_name') or '').lower())
            df_players = pd.DataFrame(players_data)
            df_players = sanitize_excel_strings(df_players)
            df_players.to_excel(xw, sheet_name='Players', index=False)

        # Format Players-related sheets
        for sheet_name in ['Players','Players_Clean','Players_Junk']:
            if sheet_name in xw.sheets:
                ws = xw.sheets[sheet_name]
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = ws.dimensions

        # Build Divisions sheet: one row per division_canon
        divisions_map = defaultdict(lambda: {"placements": 0, "events": set()})
        
        for rec in records:
            eid = rec.get("event_id")
            placements = rec.get("placements", [])
            for p in placements:
                div_canon = p.get("division_canon", "")
                div_category = p.get("division_category", "unknown")
                if div_canon:
                    divisions_map[div_canon]["placements"] += 1
                    divisions_map[div_canon]["events"].add(eid)
                    # Store category (should be consistent per division, but take first seen)
                    if "category" not in divisions_map[div_canon]:
                        divisions_map[div_canon]["category"] = div_category
        
        divisions_data = [
            {
                "division_canon": div,
                "division_category": divisions_map[div].get("category", "unknown"),
                "count_placements": divisions_map[div]["placements"],
                "count_events": len(divisions_map[div]["events"]),
            }
            for div in sorted(divisions_map.keys())
        ]
        
        df_divisions = pd.DataFrame(divisions_data)
        df_divisions = sanitize_excel_strings(df_divisions)
        df_divisions.to_excel(xw, sheet_name="Divisions", index=False)
        
        divisions_ws = xw.sheets["Divisions"]
        divisions_ws.freeze_panes = "A2"
        divisions_ws.auto_filter.ref = divisions_ws.dimensions

        # ------------------------------------------------------------
        # Placements_Flat sheet (truth-preserving analysis table)
        # ------------------------------------------------------------
        if placements_flat_df is not None and len(placements_flat_df) > 0:
            df_pf = placements_flat_df.copy()
            if "year" in df_pf.columns:
                df_pf["year"] = pd.to_numeric(df_pf["year"], errors="coerce").astype("Int64")
            df_pf = sanitize_excel_strings(df_pf)
            df_pf.to_excel(xw, sheet_name="Placements_Flat", index=False)
            ws = xw.sheets["Placements_Flat"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # ------------------------------------------------------------
        # IMPORTANT:
        # Stage 04 owns the final strict Option-A Persons_Truth (presentation/pivot-ready).
        # To avoid confusion/overwrites, Stage 03 writes its source/tracing version under
        # a distinct sheet name (no data changes).
        # ------------------------------------------------------------
        persons_truth_df = build_persons_truth(placements_flat_df)
        if persons_truth_df is not None and not persons_truth_df.empty:
            persons_truth_df = sanitize_excel_strings(persons_truth_df)
            persons_truth_df.to_excel(xw, sheet_name="Persons_Truth_Source", index=False)

            ws = xw.sheets["Persons_Truth_Source"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # ------------------------------------------------------------
        # Persons_Unresolved sheet (from out/Persons_Unresolved*.csv; triage preferred)
        # ------------------------------------------------------------
        if unresolved_df is not None and not unresolved_df.empty:
            df_ur = sanitize_excel_strings(unresolved_df.copy())
            df_ur.to_excel(xw, sheet_name="Persons_Unresolved", index=False)
            ws = xw.sheets["Persons_Unresolved"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # ------------------------------------------------------------
        # Divisions_Normalized sheet (presentation-only grouping)
        # ------------------------------------------------------------
        norm_map = defaultdict(lambda: {
            "placements": 0,
            "events": set(),
            "categories": set(),
            "aliases": set(),
        })

        for rec in records:
            eid = rec.get("event_id")
            placements = rec.get("placements", [])
            for p in placements:
                div_canon = (p.get("division_canon") or "").strip()
                div_raw = (p.get("division_raw") or "").strip()
                div_cat = (p.get("division_category") or "unknown").strip() or "unknown"

                # Use canon if present; also record raw as alias
                base = div_canon or div_raw
                if not base:
                    continue

                key = normalize_division_key(base)
                if not key:
                    continue

                nm = norm_map[key]
                nm["placements"] += 1
                if eid:
                    nm["events"].add(eid)
                if div_cat:
                    nm["categories"].add(div_cat)

                if div_canon:
                    nm["aliases"].add(div_canon)
                if div_raw:
                    nm["aliases"].add(div_raw)

        divisions_norm_data = []
        for key, nm in norm_map.items():
            aliases = sorted(nm["aliases"], key=lambda x: (len(x), x.lower()))
            # If multiple categories seen, keep 'mixed' to be honest
            cats = sorted([c for c in nm["categories"] if c])
            cat = cats[0] if len(cats) == 1 else ("mixed" if cats else "unknown")

            years_seen = sorted(
                {rec.get("year") for rec in records if rec.get("event_id") in nm["events"] and rec.get("year") is not None}
            )

            divisions_norm_data.append({
                "division_key": key,
                "division_display": pick_best_division_display(aliases),
                "division_category": cat,
                "count_placements": nm["placements"],
                "count_events": len(nm["events"]),
                "first_year_seen": years_seen[0] if years_seen else "",
                "last_year_seen": years_seen[-1] if years_seen else "",
                "aliases": " | ".join(aliases),
            })

        if divisions_norm_data:
            # Highest frequency first
            divisions_norm_data.sort(key=lambda r: (-int(r["count_placements"]), r["division_display"].lower()))
            df_div_norm = pd.DataFrame(divisions_norm_data)
            df_div_norm = sanitize_excel_strings(df_div_norm)
            df_div_norm.to_excel(xw, sheet_name="Divisions_Normalized", index=False)

            ws = xw.sheets["Divisions_Normalized"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # Build Teams sheet: team placements only
        teams_data = []
        for rec in records:
            eid = rec.get("event_id")
            year = rec.get("year")
            placements = rec.get("placements", [])
            for p in placements:
                if p.get("competitor_type") == "team":
                    division_name = p.get("division_canon", "") or p.get("division_raw", "")
                    if not is_team_division(division_name):
                        continue  # skip singles
                    p1 = sanitize_string(p.get("player1_name", ""))
                    p2 = sanitize_string(p.get("player2_name", ""))

                    teams_data.append({
                        "event_id": eid,
                        "year": year if year is not None else "",
                        "division_canon": sanitize_string(p.get("division_canon", "")),
                        "place": p.get("place", ""),
                        "team_display_name": team_display_name(p1, p2),
                        "player1_name": p1,
                        "player2_name": p2,
                        "player1_id": sanitize_string(p.get("player1_id") or p.get("player_id") or ""),
                        "player2_id": sanitize_string(p.get("player2_id") or ""),
                    })
        
        if teams_data:
            df_teams = pd.DataFrame(teams_data)
            df_teams = sanitize_excel_strings(df_teams)
            df_teams.to_excel(xw, sheet_name="Teams", index=False)
            
            teams_ws = xw.sheets["Teams"]
            teams_ws.freeze_panes = "A2"
            teams_ws.auto_filter.ref = teams_ws.dimensions

            # ------------------------------------------------------------
            # Teams_Alias_Candidates (presentation-only; NO merges)
            # Group by normalized member-name key (order-invariant).
            # ------------------------------------------------------------
            df_t = pd.DataFrame(teams_data).copy()
            for col in ["player1_name","player2_name","player1_id","player2_id","division_canon","event_id","year","place"]:
                if col not in df_t.columns:
                    df_t[col] = ""

            df_t["team_key"] = df_t.apply(lambda r: normalize_team_key(r.get("player1_name",""), r.get("player2_name","")), axis=1)
            df_t = df_t[df_t["team_key"].astype(str).str.len() > 0]

            cand_rows = []
            for tkey, g in df_t.groupby("team_key"):
                # Build alias strings like "Name1 / Name2" as seen in data
                alias_pairs = []
                for _, r in g.iterrows():
                    n1 = str(r.get("player1_name","") or "").strip()
                    n2 = str(r.get("player2_name","") or "").strip()
                    if n1 and n2:
                        alias_pairs.append(f"{n1} / {n2}")

                # Unique aliases (case-insensitive)
                seen = set()
                alias_u = []
                for a in alias_pairs:
                    k = a.lower().strip()
                    if k and k not in seen:
                        seen.add(k)
                        alias_u.append(a.strip())

                # ID pairs observed (order-invariant)
                id_pairs = set()
                for _, r in g.iterrows():
                    i1 = str(r.get("player1_id","") or "").strip()
                    i2 = str(r.get("player2_id","") or "").strip()
                    if i1 and i2:
                        left, right = sorted([i1, i2])
                        id_pairs.add(f"{left} | {right}")
                id_pairs = sorted(id_pairs)

                # Only interesting if there is some variation:
                # - more than 1 alias spelling/order OR
                # - more than 1 distinct id-pair (signals inconsistent player IDs)
                if len(alias_u) < 2 and len(id_pairs) < 2:
                    continue

                divisions = sorted(set([str(x).strip() for x in g["division_canon"].dropna().astype(str).tolist() if str(x).strip()]))
                years = sorted(set([str(x).strip() for x in g["year"].dropna().astype(str).tolist() if str(x).strip()]))

                cand_rows.append({
                    "candidate_group_id": _stable_group_id("t", tkey),
                    "team_key": tkey,
                    "team_display_best": _best_team_display(alias_u),
                    "aliases": " | ".join(alias_u),
                    "player_id_pairs": " || ".join(id_pairs),   # keep for traceability (can hide in Excel)
                    "divisions_seen": " | ".join(divisions),
                    "years_seen": " | ".join(years),
                    "count_placements": int(len(g)),
                    "confidence": "med" if len(alias_u) >= 2 else "low",
                    "decision": "",  # merge / not_merge / unsure
                    "notes": "",
                })

            if cand_rows:
                cand_rows.sort(key=lambda r: (-int(r.get("count_placements") or 0),
                                             (r.get("team_display_best") or "").lower()))
                df_tc = pd.DataFrame(cand_rows)
                df_tc = sanitize_excel_strings(df_tc)
                df_tc.to_excel(xw, sheet_name="Teams_Alias_Candidates", index=False)

                ws = xw.sheets["Teams_Alias_Candidates"]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        # Build QC_TopIssues sheet: from QC output files
        qc_summary_path = OUT_DIR / "stage3_qc_summary.json"
        qc_issues_path = OUT_DIR / "stage3_qc_issues.jsonl"
        
        if qc_summary_path.exists() and qc_issues_path.exists():
            try:
                
                
                # Read summary
                with open(qc_summary_path, "r", encoding="utf-8") as f:
                    qc_summary = json.load(f)
                
                # Build counts by check_id table
                counts_by_check = qc_summary.get("counts_by_check", {})
                qc_counts_data = []
                for check_id, counts in counts_by_check.items():
                    qc_counts_data.append({
                        "check_id": check_id,
                        "ERROR": counts.get("ERROR", 0),
                        "WARN": counts.get("WARN", 0),
                        "INFO": counts.get("INFO", 0),
                        "Total": counts.get("ERROR", 0) + counts.get("WARN", 0) + counts.get("INFO", 0),
                    })
                qc_counts_data.sort(key=lambda x: x["Total"], reverse=True)
                
                # Read sample issues
                qc_issues_data = []
                with open(qc_issues_path, "r", encoding="utf-8") as f:
                    for line_num, line in enumerate(f, 1):
                        if line_num > 100:  # Limit to first 100 issues
                            break
                        try:
                            issue = json.loads(line.strip())
                            qc_issues_data.append({
                                "event_id": issue.get("event_id", ""),
                                "check_id": issue.get("check_id", ""),
                                "severity": issue.get("severity", ""),
                                "field": issue.get("field", ""),
                                "message": sanitize_string(issue.get("message", ""))[:100],
                                "example_value": sanitize_string(str(issue.get("example_value", "")))[:50],
                                "context": sanitize_string(str(issue.get("context", "")))[:100],
                            })
                        except json.JSONDecodeError:
                            continue
                
                # Write QC sheet
                if qc_counts_data:
                    df_qc_counts = pd.DataFrame(qc_counts_data)
                    df_qc_counts = sanitize_excel_strings(df_qc_counts)
                    df_qc_counts.to_excel(xw, sheet_name="QC_TopIssues", index=False, startrow=0)
                
                if qc_issues_data:
                    df_qc_issues = pd.DataFrame(qc_issues_data)
                    df_qc_issues = sanitize_excel_strings(df_qc_issues)
                    start_row = len(qc_counts_data) + 3 if qc_counts_data else 0
                    df_qc_issues.to_excel(xw, sheet_name="QC_TopIssues", index=False, startrow=start_row)
                
                qc_ws = xw.sheets.get("QC_TopIssues")
                if qc_ws:
                    qc_ws.freeze_panes = "A2"
                    qc_ws.auto_filter.ref = qc_ws.dimensions
            except Exception as e:
                # Silently skip QC sheet if files don't exist or are malformed
                pass

        # ------------------------------------------------------------
        # README sheet (one-page explanation of workbook structure)
        # ------------------------------------------------------------
        readme_rows = [
            {
                "Section": "Purpose",
                "Description": (
                    "Canonical archive of historical footbag competition results. "
                    "Identity is locked: every placement maps to a verified person record "
                    "(Persons_Truth) or is explicitly marked unresolved. "
                    "Reproducible from source — do not edit sheets directly."
                ),
            },
            {
                "Section": "ANALYTICS — Safe for pivot tables and stats",
                "Description": (
                    "These sheets use identity-locked, coverage-filtered data and can be "
                    "trusted for analysis:\n"
                    "- Analytics_Safe_Surface: Flat placement rows, complete/mostly_complete "
                    "coverage only, identity resolved. Primary source for all stats.\n"
                    "- Person_Stats: Career stats per person (events, wins, podiums, year range) "
                    "— derived from Analytics_Safe_Surface.\n"
                    "- PersonStats_ByDivCat: Same, broken down by division category "
                    "(freestyle / net / sideline / golf).\n"
                    "- Persons_Truth: Canonical person registry — one row per verified human, "
                    "identity locked.\n"
                    "- Division_Stats: Per-division aggregates.\n"
                    "- Coverage_ByEventDiv: Coverage quality (complete / mostly_complete / "
                    "partial / sparse) for every event×division combination."
                ),
            },
            {
                "Section": "EVENT ARCHIVE — Browse by year",
                "Description": (
                    "- Year sheets (1980–present): One column per event, raw results text "
                    "plus structured placement data. Ordered chronologically within each year.\n"
                    "- Index: Full event list with hyperlinks to year sheets.\n"
                    "- Summary: Aggregate counts (events, placements, year range).\n"
                    "- Divisions / Divisions_Normalized: Canonical division labels and "
                    "near-identical groupings."
                ),
            },
            {
                "Section": "COMPLETE RECORD — Unfiltered (includes partial coverage)",
                "Description": (
                    "These sheets contain all placements, including events with incomplete "
                    "coverage. Counts will differ from Analytics_Safe_Surface.\n"
                    "- Placements_Flat: All 25,669 placement rows. "
                    "coverage_flag column indicates quality.\n"
                    "- Placements_ByPerson: Same rows with person_id and person_canon.\n"
                    "- Persons_Unresolved: 449 ambiguous/unresolved identities "
                    "(abbreviated names, single-name handles, unclear aliases).\n"
                    "- Placements_Unresolved: 155 placements whose person cannot be "
                    "fully resolved with current evidence.\n"
                    "- Data_Integrity: Pipeline integrity metrics and coverage summary."
                ),
            },
            {
                "Section": "Known Limitations",
                "Description": (
                    "- 155 placements remain unresolved (tied to 76 real-person ambiguities "
                    "such as 'Jean', 'Pierre', abbreviated last initials).\n"
                    "- 16 person entries excluded: two-person concatenated strings that "
                    "could not be split without guessing "
                    "(e.g. 'Manuel Kruse Simon Voss').\n"
                    "- Partial/sparse coverage in ~4.4% of event×division combinations — "
                    "see Coverage_ByEventDiv. These rows appear in Placements_Flat but "
                    "not in Analytics_Safe_Surface.\n"
                    "- Identity is frozen at v31/v27/v33. Any new person identifications "
                    "require a new versioned release."
                ),
            },
        ]

        # Prefer external README CSV if present (repo root); fallback to inline readme_rows
        readme_csv_candidates = [
            Path("readme-excel.csv"),
            Path("readme_excel.csv"),
        ]

        df_readme = None
        for p in readme_csv_candidates:
            if p.exists():
                df_readme = pd.read_csv(
                    p, dtype=str, quoting=csv.QUOTE_MINIMAL
                ).fillna("")
                break

        if df_readme is None:
            df_readme = pd.DataFrame(readme_rows).fillna("")

        df_readme = sanitize_excel_strings(df_readme)
        df_readme.to_excel(xw, sheet_name="README", index=False)

        ws = xw.sheets["README"]
        ws.freeze_panes = "A2"
        ncols = len(df_readme.columns)
        ws.column_dimensions["A"].width = 24
        if ncols >= 4:
            ws.column_dimensions["B"].width = 8
            ws.column_dimensions["C"].width = 40
            ws.column_dimensions["D"].width = 120
        else:
            ws.column_dimensions["B"].width = 120
        ws.auto_filter.ref = ws.dimensions
        # Wrap text in description column (last column) so full content is visible
        desc_col = ncols
        for row_idx in range(2, len(df_readme) + 2):
            cell = ws.cell(row=row_idx, column=desc_col)
            a = copy(cell.alignment)
            a.wrap_text = True
            cell.alignment = a

    # Save event_locator map for downstream hyperlinks (Stage 04)
    locator_path = OUT_DIR / "event_locator.json"
    with open(locator_path, "w", encoding="utf-8") as f:
        json.dump(event_locator, f)


def print_verification_stats(records: list[dict], out_xlsx: Path) -> None:
    """Print verification gate statistics."""
    total = len(records)
    print(f"\n{'='*60}")
    print("VERIFICATION GATE: Stage 3 (Excel Output)")
    print(f"{'='*60}")
    print(f"Total events in output: {total}")

    if total == 0:
        return

    # Count by year
    by_year = {}
    unknown = 0
    for rec in records:
        year = rec.get("year")
        if year is not None:
            by_year[year] = by_year.get(year, 0) + 1
        else:
            unknown += 1

    years = sorted(by_year.keys())
    print(f"\nSheet count: {len(years)} year sheets" + (", 1 unknown_year sheet" if unknown else ""))

    if years:
        print(f"Year range: {min(years)} - {max(years)}")

    print("\nEvents per sheet (first 10):")
    for y in years[:10]:
        print(f"  {int(y)}.0: {by_year[y]} events")
    if len(years) > 10:
        print(f"  ... and {len(years) - 10} more year sheets")
    if unknown:
        print(f"  unknown_year: {unknown} events")

    # Spot check 10 events
    print("\nSpot check (10 sample events):")
    import random
    sample = random.sample(records, min(10, len(records)))
    for rec in sample:
        eid = rec.get("event_id")
        year = rec.get("year")
        name = str(rec.get("event_name", ""))[:30]
        placements = len(rec.get("placements", []))
        print(f"  {eid:6s} | {year or '????'} | {name:30s} | {placements} placements")

    print(f"\nOutput file: {out_xlsx}")
    print(f"{'='*60}\n")


def main():
    """
    Read stage2 CSV and output final Excel workbook.
    """
    in_csv = OUT_DIR / "stage2_canonical_events.csv"
    out_xlsx = REPO_ROOT / "Footbag_Results_Canonical.xlsx"

    # Unresolved persons: prefer triage file if present (from 04 or manual)
    unresolved_triage = OUT_DIR / "Persons_Unresolved_Triage.csv"
    unresolved_base = OUT_DIR / "Persons_Unresolved.csv"
    df_unresolved = _read_optional_csv(unresolved_triage)
    if df_unresolved.empty:
        df_unresolved = _read_optional_csv(unresolved_base)

    players_csv = OUT_DIR / "stage2p5_players_clean.csv"
    players_df = None
    if players_csv.exists():
        players_df = pd.read_csv(players_csv)
    else:
        print(f"Info: No stage2p5 players file at {players_csv}; using placement-derived Players.")

    if not in_csv.exists():
        print(f"ERROR: Input file not found: {in_csv}")
        print("Run 02_canonicalize_results.py first.")
        return

    # 1. Load the two different datasets
    events_df = pd.read_csv(in_csv)
    placements_flat_csv = OUT_DIR / "Placements_Flat.csv"
    if not placements_flat_csv.exists():
        print(f"ERROR: Missing required {placements_flat_csv}. Run 02p5 to generate it.", file=sys.stderr)
        return
    placements_df = pd.read_csv(placements_flat_csv)

    # 2. Ensure IDs are the same type (both strings)
    events_df["event_id"] = events_df["event_id"].astype(str)
    placements_df["event_id"] = placements_df["event_id"].astype(str)

    # 1. Ensure we only have ONE record per event for the Excel columns
    events_df = events_df.drop_duplicates(subset=["event_id"]).copy()
    records = events_df.to_dict("records")
    print(f"Verified: Writing {len(records)} unique events to Excel.")
    print(f"Loaded {placements_flat_csv} ({len(placements_df)} rows)")

    # 2. Pre-build the results blobs from the FLAT placements file
    # This solves the "0 placements" issue
    players_by_id = build_players_by_id(players_df)
    results_map = {}
    if not placements_df.empty:
        print(f"Grouping {len(placements_df)} placements into events...")
        for eid, group in placements_df.groupby("event_id"):
            results_map[str(eid)] = format_results_from_placements(group.to_dict("records"), players_by_id)

    # 3. Pass the corrected data to the writer
    write_excel(
        out_xlsx,
        records,
        players_df=players_df,
        placements_flat_df=placements_df,
        unresolved_df=df_unresolved,
        events_df=events_df,
        results_map=results_map,
    )

    # Run Stage 3 QC on Excel workbook data
    if USE_MASTER_QC:
        qc_summary, qc_issues = qc_master.run_qc_for_stage(
            "stage3", records, results_map=results_map, players_by_id=players_by_id, out_dir=OUT_DIR
        )
        qc_master.print_qc_summary(qc_summary, "stage3")
    else:
        print("Skipping Stage 3 QC (qc_master not available)")

    print_verification_stats(records, out_xlsx)
    print(f"Wrote: {out_xlsx}")


if __name__ == "__main__":
    main()
