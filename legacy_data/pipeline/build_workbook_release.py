#!/usr/bin/env python3
"""
pipeline/build_workbook_release.py

v22-style release workbook — primary spreadsheet deliverable.

Ported from build_workbook_v17.py (FOOTBAG_DATA lineage).
Reads from event_results/canonical_input/ (platform-facing, already filtered).

Sheet order:
    README → EVENT INDEX → STATISTICS → ERA LEADERS →
    PLAYER STATS → QC - EXCLUDED EVENTS → <year sheets>

Year sheets: non-sparse events only (FULL / PARTIAL / QUARANTINED).
EVENT INDEX: all events (including sparse / no results).
Excluded: Consecutive Records, Freestyle Insights.

Run after:
    ./run_pipeline.sh rebuild && ./run_pipeline.sh release
    .venv/bin/python pipeline/qc/run_qc.py

Usage:
    .venv/bin/python pipeline/build_workbook_release.py
"""

import csv
import os
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

csv.field_size_limit(sys.maxsize)

ROOT = Path(__file__).resolve().parents[1]

# ── Paths ──────────────────────────────────────────────────────────────────────
CANONICAL_INPUT    = ROOT / "event_results" / "canonical_input"
CANONICAL_UPSTREAM = ROOT / "out" / "canonical"   # for early-era person supplement
QUARANTINE_CSV     = ROOT / "inputs" / "review_quarantine_events.csv"
OUTPUT_PATH        = ROOT / "out" / "Footbag_Results_Release.xlsx"

# Threshold for early-era person supplement (section 3 of alignment audit).
# Persons with first_year ≤ this value are included even without placements.
_EARLY_ERA_FIRST_YEAR_CUTOFF = 1990

VERSION = "v22"
UPDATED = date.today().isoformat()

# ── Styles ─────────────────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

FILL_NONE    = PatternFill(fill_type=None)
FILL_HEADER  = _fill("D9D9D9")
FILL_SECTION = _fill("1F3864")
FILL_WORLDS  = _fill("17375E")
FILL_WORLDS_ROW = _fill("C9DCF0")
FILL_TITLE   = _fill("2E75B6")
FILL_NET     = _fill("E2EFDA")
FILL_FREE    = _fill("FFF2CC")
FILL_GOLF    = _fill("FCE4D6")
FILL_OTHER   = _fill("F2F2F2")
FILL_EVENT_A = _fill("DEEAF1")
FILL_EVENT_B = _fill("FFFFFF")
FILL_HOF     = _fill("FFF2CC")
FILL_BAP     = _fill("E2EFDA")
FILL_BOTH    = _fill("EAD1F5")
FILL_QUAR    = _fill("FFE0E0")
FILL_WARN    = _fill("FFF2CC")
WHITE        = _fill("FFFFFF")

FONT_TITLE   = Font(bold=True,  size=13, color="FFFFFF")
FONT_SECTION = Font(bold=True,  size=11, color="FFFFFF")
FONT_HEADER  = Font(bold=True,  size=11)
FONT_DATA    = Font(size=11)
FONT_SMALL   = Font(size=9,    color="808080")
FONT_NOTE    = Font(size=9,    italic=True, color="606060")
FONT_LINK    = Font(size=11,   color="0563C1", underline="single")
FONT_UNKNOWN = Font(size=10,   italic=True, color="A0A0A0")

ALIGN_L   = Alignment(horizontal="left",   vertical="top", wrap_text=False)
ALIGN_R   = Alignment(horizontal="right",  vertical="top")
ALIGN_C   = Alignment(horizontal="center", vertical="top")
ALIGN_LW  = Alignment(horizontal="left",   vertical="top", wrap_text=True)

# ── Sentinels (never included in stats / leaderboards) ─────────────────────────
_SKIP_PID   = {"", "__NON_PERSON__"}
_SKIP_DNAME = {"__NON_PERSON__", "__UNKNOWN_PARTNER__", "[UNKNOWN PARTNER]", "[UNKNOWN]", ""}
_UNKNOWN_DISPLAY = "[Unknown]"

# ── Helpers ─────────────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font   is not None: cell.font          = font
    if fill   is not None: cell.fill          = fill
    if align  is not None: cell.alignment     = align
    if number_format:      cell.number_format = number_format
    return cell


def _title_row(ws, row: int, text: str, ncols: int = 8) -> int:
    _w(ws, row, 1, text, font=FONT_TITLE, fill=FILL_TITLE, align=ALIGN_L)
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = FILL_TITLE
    ws.row_dimensions[row].height = 22
    return row + 1


def _section_row(ws, row: int, text: str, ncols: int = 8,
                 fill=None, font=None) -> int:
    f  = fill or FILL_SECTION
    fn = font or FONT_SECTION
    _w(ws, row, 1, text, font=fn, fill=f, align=ALIGN_L)
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = f
    ws.row_dimensions[row].height = 18
    return row + 1


def _worlds_section_row(ws, row: int, text: str, ncols: int = 8) -> int:
    return _section_row(ws, row, text, ncols=ncols, fill=FILL_WORLDS)


def _hrow(ws, row: int, *headers) -> int:
    for col, h in enumerate(headers, 1):
        _w(ws, row, col, h, font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
    return row + 1


def _drow(ws, row: int, *values, fill=None) -> int:
    for col, v in enumerate(values, 1):
        align = ALIGN_R if isinstance(v, (int, float)) else ALIGN_L
        cell  = _w(ws, row, col, v, font=FONT_DATA, align=align)
        if fill:
            cell.fill = fill
    return row + 1


def _note_row(ws, row: int, text: str, ncols: int = 5) -> int:
    _w(ws, row, 1, text, font=FONT_NOTE, align=ALIGN_L)
    ws.row_dimensions[row].height = 13
    return row + 1


def _load(name: str) -> list[dict]:
    path = CANONICAL_INPUT / name
    if not path.exists():
        print(f"  [WARN] missing {path}")
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _cat_fill(cat: str) -> PatternFill:
    c = (cat or "").lower()
    if c == "net":       return FILL_NET
    if c == "freestyle": return FILL_FREE
    if c in ("golf", "sideline"): return FILL_GOLF
    return FILL_OTHER


def _location(ev: dict) -> str:
    """Construct a display location string from city/region/country."""
    parts = [ev.get("city", ""), ev.get("region", ""), ev.get("country", "")]
    return ", ".join(p for p in parts if p)


def _is_worlds(ev: dict) -> bool:
    """True for World Championship events (post-1997 event_type, or pre-1997 by key)."""
    if ev.get("event_type", "") == "worlds":
        return True
    return "worlds" in ev.get("event_key", "").lower()


# ── Worlds naming normalization (presentation layer only) ──────────────────────
# Rule: display name = "{N}th Annual World Footbag Championships"
#       where N = year - 1979.
# Applied for years ≥ 1985. Pre-1984 has multiple competing org events per year
# and is left with raw names pending authoritative merge decisions.
#
# The year-based formula (N = year - 1979) is intentional: it counts from the
# first championship year (1980 = 1st) and is year-indexed, not held-count.
# COVID-skipped years (2020/2021) are absorbed into the number, matching v22.

def _ordinal(n: int) -> str:
    """Return ordinal string: 1 → '1st', 2 → '2nd', 3 → '3rd', 4 → '4th', etc."""
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    suffix = ["th", "st", "nd", "rd"] + ["th"] * 6
    return f"{n}{suffix[n % 10]}"


def _worlds_display_name(year: int) -> str:
    """Return normalized worlds event display name for a given year."""
    n = year - 1979
    return f"{_ordinal(n)} Annual World Footbag Championships"


def _worlds_name_override(ev: dict) -> str | None:
    """Return normalized display name if this is a worlds event with year ≥ 1984,
    otherwise None (keep raw name).

    Post-1997 events with event_type="worlds" are always overridden.
    Pre-1997 events use a name-content guard to exclude false positives where
    "worlds" appears in the event_key incidentally (e.g. "Worlds Warm-Up" events).
    Pre-1984: multiple org events per year — do not auto-rename.
    """
    try:
        year = int(ev.get("year", "") or 0)
    except ValueError:
        return None
    if year <= 1984:
        return None

    etype = ev.get("event_type", "")

    # Post-1997 standard: event_type == "worlds" is the authoritative signal.
    if etype == "worlds":
        return _worlds_display_name(year)

    # Pre-1997: "worlds" in event_key but event_type is a sport category.
    # Require the event_name to contain a worlds-championship signal to avoid
    # false positives (e.g. 1997_worlds_vancouver = "Vancouver Open (Worlds Warm-Up)").
    if "worlds" not in ev.get("event_key", "").lower():
        return None
    name_lower = ev.get("event_name", "").lower()
    _WORLDS_SIGNALS = (
        "world footbag", "world championship", "wfa world",
        "ifab world", "nhsa world", "ifpa world",
    )
    if not any(sig in name_lower for sig in _WORLDS_SIGNALS):
        return None
    return _worlds_display_name(year)


# ── Canonical ordering for year/event sheet layout ─────────────────────────────

_CAT_ORDER: dict[str, int] = {
    "OVERALL":     0,
    "CONSECUTIVE": 1,
    "NET":         2,
    "FREESTYLE":   3,
    "GOLF":        4,
    "DISTANCE":    5,
    "ACCURACY":    6,
    "SIDELINE":    7,
    "UNKNOWN":     8,
}

_DIV_ORDER: dict[str, dict[str, int]] = {
    "OVERALL": {
        "Men's Overall":              0,
        "Open Overall":               1,
        "Women's Overall":            2,
        "Freestyle Overall":          3,
        "Intermediate Overall":       4,
        "Novice Overall":             5,
    },
    "CONSECUTIVE": {
        "Open Singles Consecutive":       0,
        "Open Doubles Consecutive":       1,
        "Women's Singles Consecutive":    2,
        "Women's Doubles Consecutive":    3,
        "Intermediate Singles Consecutive": 4,
        "Singles Consecutive":            5,
        "Doubles Consecutive":            6,
    },
    "NET": {
        "Open Singles Net":           0,
        "Open Doubles Net":           1,
        "Open Mixed Doubles Net":     2,
        "Mixed Doubles Net":          3,
        "Women's Singles Net":        4,
        "Women's Doubles Net":        5,
        "Intermediate Singles Net":   6,
        "Intermediate Doubles Net":   7,
    },
    "FREESTYLE": {
        "Open Singles Freestyle":         0,
        "Open Singles Routines":          1,
        "Women's Singles Freestyle":      2,
        "Women's Singles Routines":       3,
        "Intermediate Singles Freestyle": 4,
        "Open Doubles Freestyle":         5,
        "Open Team Freestyle":            6,
        "Mixed Doubles Freestyle":        7,
    },
    "GOLF": {
        "Open Golf":                  0,
        "Open Singles Golf":          1,
        "Open Doubles Golf":          2,
        "Women's Golf":               3,
        "Intermediate Golf":          4,
        "Novice Golf":                5,
    },
}


def _disc_sort_key(disc_name: str, disc_rec: dict) -> tuple:
    """Sort key: canonical category order → canonical division order → alpha fallback.
    discipline_category in canonical_input is lowercase; uppercase for _CAT_ORDER lookup.
    """
    cat = (disc_rec.get("discipline_category") or "").upper()
    div = disc_rec.get("discipline_name") or disc_name
    cat_idx = _CAT_ORDER.get(cat, 99)
    div_idx = _DIV_ORDER.get(cat, {}).get(div, 999)
    return (cat_idx, div_idx, div.lower(), disc_name.lower())


def _dedup_slot(slot_rows: list) -> list:
    """Deduplicate participant rows by person_id, preserving order."""
    seen: set = set()
    result: list = []
    for row in slot_rows:
        pid = row.get("person_id", "").strip()
        key = pid if pid else id(row)
        if key not in seen:
            seen.add(key)
            result.append(row)
    return result


# ── Ranking helper ─────────────────────────────────────────────────────────────

def _add_ranks(rows: list, sort_key_idx: int) -> list:
    """Prepend a rank string with T-N tie notation. Rows pre-sorted descending."""
    if not rows:
        return []
    from collections import Counter as C
    val_count: dict = C(r[sort_key_idx] for r in rows)

    result = []
    rank = 1
    prev_val = None
    run_start = 1
    for i, row in enumerate(rows, 1):
        val = row[sort_key_idx]
        if val != prev_val:
            run_start = rank
        rank_str = f"T-{run_start}" if val_count[val] > 1 else str(run_start)
        result.append((rank_str,) + tuple(row))
        if val != prev_val:
            prev_val = val
        if i < len(rows) and rows[i][sort_key_idx] != val:
            rank = i + 1

    return result


# ── Data loading ───────────────────────────────────────────────────────────────

def load_all():
    print("Loading canonical_input CSVs…")

    raw_events  = _load("events.csv")
    raw_discs   = _load("event_disciplines.csv")
    raw_results = _load("event_result_participants.csv")
    raw_persons = _load("persons.csv")

    events:  dict[str, dict] = {r["event_key"]: r for r in raw_events}
    discs:   dict[tuple, dict] = {}
    for r in raw_discs:
        discs[(r["event_key"], r["discipline_key"])] = r
    persons: dict[str, dict] = {r["person_id"]: r for r in raw_persons}

    # Quarantine: review_quarantine_events.csv uses legacy integer event_ids.
    # Map via legacy_event_id column in events.csv → event_key.
    legacy_id_to_key: dict[str, str] = {
        r["legacy_event_id"]: r["event_key"]
        for r in raw_events
        if r.get("legacy_event_id", "").strip()
    }
    quarantine: set[str] = set()
    if QUARANTINE_CSV.exists():
        with open(QUARANTINE_CSV, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                eid_str = r.get("event_id", "").strip()
                if eid_str in legacy_id_to_key:
                    quarantine.add(legacy_id_to_key[eid_str])
                elif eid_str and eid_str in events:
                    quarantine.add(eid_str)

    # ── Early-era person supplement ──────────────────────────────────────────
    # Persons with first_year ≤ _EARLY_ERA_FIRST_YEAR_CUTOFF in the upstream
    # canonical persons file are included in the workbook even if they have no
    # surviving participant rows, no member_id, and no BAP/HOF — because their
    # presence is documented in the identity lock from authoritative sources
    # (magazines, TXT files).  This is a workbook-only supplement; it does not
    # modify canonical_input or the platform export.
    #
    # Person-likeness + alias-dedup gates are applied so the supplement doesn't
    # re-introduce junk or duplicate entries that the platform export filtered.
    import re as _re
    _PL_MOJIBAKE   = _re.compile(r"[¶¦±¼¿¸¹º³]")
    _PL_EMBED_Q    = _re.compile(r"\w\?|\?\w")
    _PL_STANDALONE = _re.compile(r"(?:^|\s)\?{1,5}(?:\s|$)")
    _PL_BAD_CHARS  = _re.compile(r"[+=\\|/]")
    _PL_ABBREVIATED = _re.compile(r"^[A-Z]\.?\s+\S")
    _PL_INCOMPLETE = _re.compile(r"^\S+\s+[A-Z]$")

    def _supp_norm(name: str) -> str:
        nfkd = unicodedata.normalize("NFKD", name)
        stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
        return _re.sub(r"\s+", " ", stripped.lower().strip().replace(".", ""))

    # Load alias names to skip duplicates
    _alias_norms: set[str] = set()
    _alias_csv_path = ROOT / "overrides" / "person_aliases.csv"
    if _alias_csv_path.exists():
        with open(_alias_csv_path, newline="", encoding="utf-8") as _af:
            for _ar in csv.DictReader(_af):
                _a = _ar.get("alias", "").strip()
                if _a:
                    _alias_norms.add(_supp_norm(_a))

    # Build norm index of existing persons (already in canonical_input)
    _existing_norms: set[str] = set()
    for _p in persons.values():
        _existing_norms.add(_supp_norm(_p.get("person_name", "")))

    def _supp_is_person_like(name: str) -> bool:
        s = name.strip()
        if not s: return False
        if _PL_MOJIBAKE.search(s): return False
        if _PL_EMBED_Q.search(s): return False
        if _PL_STANDALONE.search(s): return False
        if _PL_BAD_CHARS.search(s): return False
        if "," in s: return False
        if " " not in s and "." not in s: return False
        if _PL_ABBREVIATED.match(s): return False
        if _PL_INCOMPLETE.match(s): return False
        if s[0].islower(): return False
        if _re.search(r"\bThe\b", s): return False
        if '"' in s: return False
        if " or " in s.lower(): return False
        if _re.search(r"[>]|\s:\s", s): return False
        if _re.search(r"\S{21,}", s): return False
        return True

    upstream_persons_path = CANONICAL_UPSTREAM / "persons.csv"
    if upstream_persons_path.exists():
        upstream_field = "fbhof_member"   # canonical uses fbhof_member not hof_member
        with open(upstream_persons_path, newline="", encoding="utf-8") as f:
            n_added = 0
            n_gate_skipped = 0
            n_alias_skipped = 0
            for r in csv.DictReader(f):
                pid = r.get("person_id", "").strip()
                if not pid or pid in persons:
                    continue  # already present
                pname = r.get("person_name", "").strip()
                fy_str = r.get("first_year", "").strip()
                if not fy_str:
                    continue  # no documented first_year — skip
                try:
                    fy = int(fy_str)
                except ValueError:
                    continue
                if fy > _EARLY_ERA_FIRST_YEAR_CUTOFF:
                    continue
                # Gate: skip non-person-like names
                if not _supp_is_person_like(pname):
                    n_gate_skipped += 1
                    continue
                # Alias dedup: skip if name matches a known alias or existing person
                pnorm = _supp_norm(pname)
                if pnorm in _alias_norms or pnorm in _existing_norms:
                    n_alias_skipped += 1
                    continue
                # Translate canonical field names to canonical_input field names
                persons[pid] = {
                    "person_id":          pid,
                    "person_name":        pname,
                    "member_id":          r.get("member_id", ""),
                    "country":            r.get("country", ""),
                    "first_year":         fy_str,
                    "last_year":          r.get("last_year", ""),
                    "event_count":        r.get("event_count", "0"),
                    "placement_count":    r.get("placement_count", "0"),
                    "bap_member":         r.get("bap_member", "0"),
                    "bap_nickname":       r.get("bap_nickname", ""),
                    "bap_induction_year": r.get("bap_induction_year", ""),
                    "hof_member":         r.get(upstream_field, "0"),
                    "hof_induction_year": r.get("fbhof_induction_year", r.get("hof_induction_year", "")),
                }
                _existing_norms.add(pnorm)
                n_added += 1
        if n_added or n_gate_skipped or n_alias_skipped:
            print(f"  Early-era supplement: +{n_added} persons "
                  f"(first_year ≤ {_EARLY_ERA_FIRST_YEAR_CUTOFF})"
                  f"{f', {n_gate_skipped} gate-skipped' if n_gate_skipped else ''}"
                  f"{f', {n_alias_skipped} alias-dedup-skipped' if n_alias_skipped else ''}")

    print(f"  Events: {len(events)}   Disciplines: {len(discs)}")
    print(f"  Participants: {len(raw_results)}   Persons: {len(persons)}")
    print(f"  Quarantined events: {len(quarantine)}")
    return events, discs, raw_results, persons, quarantine


def compute_pub_eids(events: dict, discs: dict, raw_results: list,
                     quarantine: set) -> set[str]:
    """Publication set for year sheets: non-sparse events with actual placement data.

    FULL / PARTIAL / QUARANTINED events are included.
    SPARSE (< 10 placements AND < 2 disciplines) and NO RESULTS are excluded.
    EVENT INDEX always shows all events regardless of this set.
    """
    plc_count: Counter = Counter(
        r["event_key"] for r in raw_results
        if r.get("person_id", "").strip() not in _SKIP_PID
    )
    disc_count: Counter = Counter(k[0] for k in discs)

    pub: set[str] = set()
    for eid, ev in events.items():
        n_plc   = plc_count.get(eid, 0)
        n_discs = disc_count.get(eid, 0)
        is_quar = eid in quarantine
        cov = _coverage_level(ev, n_discs, n_plc, is_quar)
        if cov not in ("SPARSE", "NO RESULTS"):
            pub.add(eid)
    return pub


# ── Statistics engine ──────────────────────────────────────────────────────────

def compute_stats(raw_results, events, discs, persons):
    stats: dict[str, dict] = defaultdict(lambda: {
        "wins": 0, "p1": 0, "p2": 0, "p3": 0, "podiums": 0,
        "events": set(), "years": set(),
        "worlds_wins": 0, "worlds_podiums": 0, "worlds_events": set(),
        "cat_wins":    defaultdict(int),
        "cat_podiums": defaultdict(int),
    })

    for row in raw_results:
        pid = row.get("person_id", "").strip()
        if not pid or pid in _SKIP_PID:
            continue
        dname = row.get("display_name", "").strip()
        if dname in _SKIP_DNAME:
            continue
        porder = row.get("participant_order", "1").strip()

        eid      = row.get("event_key", "").strip()
        ev       = events.get(eid, {})
        year     = ev.get("year", "").strip()
        disc_key = row.get("discipline_key", "").strip()
        disc     = discs.get((eid, disc_key), {})
        cat      = disc.get("discipline_category", "").lower().strip()

        try:
            place = int(row.get("placement", "0") or 0)
        except ValueError:
            place = 0

        s = stats[pid]
        if eid:
            s["events"].add(eid)
        if year:
            try:
                s["years"].add(int(year))
            except ValueError:
                pass

        if porder == "1":
            if place == 1:
                s["wins"] += 1
                s["p1"]   += 1
                s["cat_wins"][cat] += 1
            if 1 <= place <= 3:
                s["podiums"] += 1
                s["cat_podiums"][cat] += 1
            if place == 2: s["p2"] += 1
            if place == 3: s["p3"] += 1

            if _is_worlds(ev):
                if eid:
                    s["worlds_events"].add(eid)
                if place == 1:
                    s["worlds_wins"] += 1
                if 1 <= place <= 3:
                    s["worlds_podiums"] += 1

    result = {}
    for pid, s in stats.items():
        if pid not in persons:
            continue
        result[pid] = {
            "wins":           s["wins"],
            "p1": s["p1"], "p2": s["p2"], "p3": s["p3"],
            "podiums":        s["podiums"],
            "events":         len(s["events"]),
            "year_first":     min(s["years"]) if s["years"] else None,
            "year_last":      max(s["years"]) if s["years"] else None,
            "worlds_wins":    s["worlds_wins"],
            "worlds_podiums": s["worlds_podiums"],
            "worlds_events":  len(s["worlds_events"]),
            "cat_wins":       dict(s["cat_wins"]),
            "cat_podiums":    dict(s["cat_podiums"]),
        }
    return result


# ── README sheet ──────────────────────────────────────────────────────────────

def build_readme(wb: Workbook, events: dict, persons: dict, n_parts: int) -> None:
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 90
    ws.freeze_panes = "B2"

    n_years   = len(set(e.get("year", "") for e in events.values() if e.get("year")))
    n_events  = len(events)
    n_persons = sum(1 for p in persons.values()
                    if p.get("person_id") and p.get("person_name"))
    n_with_results = sum(1 for e in events.values()
                         if e.get("status", "") not in ("no_results", ""))

    row = _title_row(ws, 1,
        f"FOOTBAG COMPETITION RESULTS — CANONICAL RELEASE  "
        f"({VERSION} · Updated {UPDATED})", ncols=1)
    row += 1

    sections = [
        ("OVERVIEW", [
            ("Coverage",
             f"{n_with_results} events with results across {n_years} years · "
             f"{n_persons} identified competitors · {n_parts:,} placement records"),
            ("Years",    "1980 – present (all years included, no year suppression)"),
            ("Sources",  "Post-1997: Footbag.org HTML archive (mirror-derived, highest authority).  "
                         "Pre-1997: Footbag World magazine scans + oldresults.txt + "
                         "expert corrections from authoritative human sources."),
            ("Identity", "All statistics aggregated by person_id (UUID).  "
                         "No display-name matching — each person counted once regardless of "
                         "name variants, nicknames, or married-name changes.  "
                         "Canonical names sourced from persons.csv."),
            ("Divisions","Fully normalized (no Sgls/Dbls abbreviations).  "
                         "Categories: freestyle · net · golf · sideline."),
            ("Version",  f"{VERSION} — {UPDATED}"),
        ]),
        ("SHEETS", [
            ("README",               "This sheet — dataset overview, notes, known issues."),
            ("EVENT INDEX",          "All events (including sparse/excluded) with metadata."),
            ("STATISTICS",           "Career leaderboards: top 25 per category with rank and tie notation."),
            ("ERA LEADERS",          "Decade leaderboards (1980s – 2020s): top 10 podiums and wins per era."),
            ("PLAYER STATS",         "One row per identified competitor — career summary."),
            ("QC - EXCLUDED EVENTS", "Events excluded from year sheets (sparse or no results)."),
            ("<year>",               "One sheet per year (1980 – present).  Each column is one event; "
                                     "rows show placements by division.  Non-sparse events only."),
        ]),
        ("COVERAGE LIMITATIONS", [
            ("Pre-1997",        "Data reconstructed from magazine scans and text archives.  "
                                "Coverage is incomplete: some events have only partial results, "
                                "and some early events are not represented at all.  "
                                "Coverage level noted in EVENT INDEX column 'Coverage'."),
            ("Post-1997",       "Mirror-derived results are the highest-authority source.  "
                                "Most post-1997 events are complete.  Some events have partial "
                                "coverage where the source page was incomplete at time of archiving."),
            ("NHSA vs WFA",     "1980–1985 featured distinct NHSA and WFA championships; "
                                "both are represented as separate events where data permits."),
            ("Sparse events",   "Events flagged SPARSE in QC - EXCLUDED EVENTS have fewer than 3 "
                                "disciplines or fewer than 10 placements — excluded from year sheets "
                                "but listed in EVENT INDEX and QC - EXCLUDED EVENTS."),
        ]),
        ("IDENTITY MODEL", [
            ("person_id",       "Every competitor is assigned a UUID (person_id).  "
                                "Statistics count each person_id once per placement slot."),
            ("Unknowns",        "Unresolved participants display as [Unknown] in year sheets.  "
                                "These are excluded from all leaderboards and statistics."),
            ("Sentinels",       "Team slots with an unknown second member are marked [UNKNOWN PARTNER].  "
                                "Also excluded from statistics."),
            ("Name variants",   "Nickname and married-name variants are resolved to a single "
                                "canonical name via the persons table.  No duplicates."),
        ]),
        ("STATISTICS NOTES", [
            ("Worlds types",    "Worlds leaderboards include all events with event_type=worlds "
                                "plus pre-1997 events with 'worlds' in the event key "
                                "(NHSA/WFA/IFAB championships)."),
            ("Counting method", "Wins and podiums counted once per placement slot "
                                "(participant_order=1).  Team events: both members credited "
                                "equally; the slot is counted once in aggregate tables."),
            ("Top 25",          "Leaderboard tables show top 25.  Full data available "
                                "in PLAYER STATS sheet and canonical CSV files."),
            ("Ties",            "Equal-ranked positions use T-N notation (e.g. T-5).  "
                                "After a tie of N, the next rank skips N-1 positions."),
        ]),
        ("KNOWN ISSUES", [
            ("Pre-1997 gaps",   "Not all pre-1997 events are fully captured.  "
                                "Some divisions and placements below 3rd may be missing."),
            ("Unresolved IDs",  "Some participants have no resolved person_id.  "
                                "These do not affect statistics (excluded by design) "
                                "but represent genuine historical uncertainty."),
            ("Source conflicts", "Where FBW and OLD_RESULTS sources disagree, "
                                 "OLD_RESULTS (authoritative human records) takes precedence."),
        ]),
    ]

    for section_name, items in sections:
        row = _section_row(ws, row, section_name, ncols=1)
        for label, text in items:
            _w(ws, row, 1,
               f"▸ {label}:  {text}" if label else text,
               font=FONT_DATA, align=ALIGN_LW)
            ws.row_dimensions[row].height = max(16, min(60, len(text) // 6))
            row += 1
        row += 1

    ws.sheet_view.showGridLines = False
    print("  README done")


# ── STATISTICS sheet ──────────────────────────────────────────────────────────

def build_statistics(wb: Workbook, stats: dict, persons: dict) -> None:
    ws = wb.create_sheet("STATISTICS")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "STATISTICS — ALL EVENTS ALL DIVISIONS", ncols=6)

    def canon(pid):
        return persons.get(pid, {}).get("person_name", pid)

    TOP = 25

    def table(title, headers, raw_data, start_row,
              sort_key_idx=-1, is_worlds=False):
        ncols = len(headers)
        if is_worlds:
            r = _worlds_section_row(ws, start_row, title, ncols=ncols)
        else:
            r = _section_row(ws, start_row, title, ncols=ncols)
        _note_row(ws, r, f"  Top {TOP} shown · full data in PLAYER STATS sheet")
        r += 1
        r = _hrow(ws, r, *headers)
        ranked = _add_ranks(raw_data[:TOP], sort_key_idx)
        for i, row_vals in enumerate(ranked):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            r = _drow(ws, r, *row_vals, fill=fill)
        return r + 1

    row = _section_row(ws, row, "GENERAL STATISTICS", ncols=6)
    row += 1

    podium_data = sorted(
        [(canon(pid), s["p1"], s["p2"], s["p3"], s["podiums"])
         for pid, s in stats.items() if s["podiums"] > 0],
        key=lambda x: (-x[4], x[0].lower())
    )
    row = table("MOST CAREER PODIUMS (ALL DIVISIONS)",
                ["Rank", "Player", "1st", "2nd", "3rd", "Total"],
                podium_data, row, sort_key_idx=4)

    wins_data = sorted(
        [(canon(pid), s["wins"]) for pid, s in stats.items() if s["wins"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST CAREER WINS (ALL DIVISIONS)",
                ["Rank", "Player", "Wins"],
                wins_data, row, sort_key_idx=1)

    free_pod = sorted(
        [(canon(pid), s["cat_podiums"].get("freestyle", 0))
         for pid, s in stats.items() if s["cat_podiums"].get("freestyle", 0) > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST FREESTYLE PODIUMS",
                ["Rank", "Player", "Podiums"],
                free_pod, row, sort_key_idx=1)

    net_pod = sorted(
        [(canon(pid), s["cat_podiums"].get("net", 0))
         for pid, s in stats.items() if s["cat_podiums"].get("net", 0) > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST NET PODIUMS",
                ["Rank", "Player", "Podiums"],
                net_pod, row, sort_key_idx=1)

    events_data = sorted(
        [(canon(pid), s["events"]) for pid, s in stats.items() if s["events"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST EVENTS COMPETED",
                ["Rank", "Player", "Events"],
                events_data, row, sort_key_idx=1)

    career_data = sorted(
        [(canon(pid), s["year_first"], s["year_last"],
          (s["year_last"] - s["year_first"]) if s["year_first"] and s["year_last"] else 0)
         for pid, s in stats.items()
         if s["year_first"] and s["year_last"] and s["year_last"] > s["year_first"]],
        key=lambda x: (-x[3], x[0].lower())
    )
    row = table("LONGEST COMPETITIVE CAREERS",
                ["Rank", "Player", "First Year", "Last Year", "Span (yrs)"],
                career_data, row, sort_key_idx=3)

    row += 1
    row = _worlds_section_row(ws, row,
        "═══  WORLD CHAMPIONSHIPS  ═══  "
        "(event_type=worlds + pre-1997 worlds by event key)",
        ncols=6)
    row += 1

    worlds_pod = sorted(
        [(canon(pid), s["worlds_wins"], s["worlds_podiums"], s["worlds_events"])
         for pid, s in stats.items() if s["worlds_podiums"] > 0],
        key=lambda x: (-x[2], x[0].lower())
    )
    row = table("WORLDS PODIUMS",
                ["Rank", "Player", "Worlds Wins", "Worlds Podiums", "Worlds Events"],
                worlds_pod, row, sort_key_idx=2, is_worlds=True)

    worlds_wins = sorted(
        [(canon(pid), s["worlds_wins"])
         for pid, s in stats.items() if s["worlds_wins"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("WORLDS WINS",
                ["Rank", "Player", "Wins"],
                worlds_wins, row, sort_key_idx=1, is_worlds=True)

    print("  STATISTICS done")


# ── ERA LEADERS sheet ──────────────────────────────────────────────────────────

def build_era_leaders(wb: Workbook, raw_results, events, discs, persons) -> None:
    ws = wb.create_sheet("ERA LEADERS")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "ERA LEADERS — BY DECADE", ncols=6)

    _ERAS = [
        ("1980s", 1980, 1989),
        ("1990s", 1990, 1999),
        ("2000s", 2000, 2009),
        ("2010s", 2010, 2019),
        ("2020s", 2020, 2029),
    ]
    TOP = 10

    def canon(pid):
        return persons.get(pid, {}).get("person_name", pid)

    for era_name, yr_lo, yr_hi in _ERAS:
        era_podiums: dict[str, list] = defaultdict(lambda: [0, 0, 0])
        era_events:  dict[str, set]  = defaultdict(set)

        for row_r in raw_results:
            pid = row_r.get("person_id", "").strip()
            if not pid or pid in _SKIP_PID:
                continue
            if row_r.get("display_name", "").strip() in _SKIP_DNAME:
                continue
            if row_r.get("participant_order", "1").strip() != "1":
                continue

            eid = row_r.get("event_key", "").strip()
            ev  = events.get(eid, {})
            try:
                yr = int(ev.get("year", "0") or 0)
            except ValueError:
                yr = 0
            if not (yr_lo <= yr <= yr_hi):
                continue

            try:
                place = int(row_r.get("placement", "0") or 0)
            except ValueError:
                place = 0

            if 1 <= place <= 3:
                era_podiums[pid][place - 1] += 1
            if eid:
                era_events[pid].add(eid)

        row = _section_row(ws, row, f"ERA: {era_name}  ({yr_lo}–{yr_hi})", ncols=6)

        _w(ws, row, 1, "Top Podiums", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
        for c in range(2, 7):
            ws.cell(row=row, column=c).fill = FILL_HEADER
        row += 1
        row = _hrow(ws, row, "Rank", "Player", "1st", "2nd", "3rd", "Total")

        pod_rows = sorted(
            [(canon(pid), p[0], p[1], p[2], sum(p))
             for pid, p in era_podiums.items() if pid in persons and sum(p) > 0],
            key=lambda x: (-x[4], x[0].lower())
        )
        ranked_pod = _add_ranks(pod_rows[:TOP], sort_key_idx=4)
        for i, rv in enumerate(ranked_pod):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            row = _drow(ws, row, *rv, fill=fill)

        row += 1

        _w(ws, row, 1, "Top Wins", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
        for c in range(2, 4):
            ws.cell(row=row, column=c).fill = FILL_HEADER
        row += 1
        row = _hrow(ws, row, "Rank", "Player", "Wins")

        win_rows = sorted(
            [(canon(pid), p[0]) for pid, p in era_podiums.items()
             if pid in persons and p[0] > 0],
            key=lambda x: (-x[1], x[0].lower())
        )
        ranked_win = _add_ranks(win_rows[:TOP], sort_key_idx=1)
        for i, rv in enumerate(ranked_win):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            row = _drow(ws, row, *rv, fill=fill)

        row += 2

    print("  ERA LEADERS done")


# ── PLAYER STATS sheet ─────────────────────────────────────────────────────────

def build_player_stats(wb: Workbook, stats: dict, persons: dict) -> None:
    ws = wb.create_sheet("PLAYER STATS")

    headers = ["Player", "Nickname", "Country", "Events", "Wins", "Podiums", "BAP", "HOF"]
    widths  = [34, 22, 18, 9, 9, 9, 6, 6]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "PLAYER STATS — ALL IDENTIFIED COMPETITORS", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    rows_out = []
    skipped_no_stats = 0
    for pid, p in sorted(persons.items(),
                         key=lambda kv: kv[1].get("person_name", "").lower()):
        if not p.get("person_name"):
            continue
        s   = stats.get(pid, {})
        # Skip persons with no events — these are stub/phantom entries
        # (auto-generated by the seed builder for unresolved participant names)
        if s.get("events", 0) == 0 and not (
            p.get("bap_member", "") in ("1", "True", "true") or
            p.get("hof_member", "") in ("1", "True", "true")
        ):
            skipped_no_stats += 1
            continue
        bap = "Y" if p.get("bap_member", "") in ("1", "True", "true") else ""
        hof = "Y" if p.get("hof_member", "") in ("1", "True", "true") else ""
        fill = FILL_BOTH if bap and hof else (FILL_HOF if hof else (FILL_BAP if bap else None))
        rows_out.append((
            p["person_name"],
            p.get("bap_nickname", ""),
            p.get("country", ""),
            s.get("events",   0),
            s.get("wins",     0),
            s.get("podiums",  0),
            bap, hof,
            fill,
        ))

    for rec in rows_out:
        *vals, fill = rec
        r = row
        for col, v in enumerate(vals, 1):
            cell = _w(ws, r, col, v, font=FONT_DATA,
                      align=(ALIGN_R if isinstance(v, int) else ALIGN_L))
            if fill:
                cell.fill = fill
        row += 1

    if skipped_no_stats:
        print(f"  PLAYER STATS: {len(rows_out)} rows ({skipped_no_stats} stub persons with 0 events excluded)")
    else:
        print(f"  PLAYER STATS: {len(rows_out)} rows")


# ── Coverage level ─────────────────────────────────────────────────────────────

def _coverage_level(ev: dict, n_discs: int, n_plc: int, is_quar: bool) -> str:
    if is_quar:
        return "QUARANTINED"
    if ev.get("status", "") == "no_results" or n_plc == 0:
        return "NO RESULTS"
    if n_plc >= 20 and n_discs >= 3:
        return "FULL"
    if n_plc >= 10 or n_discs >= 2:
        return "PARTIAL"
    if n_plc > 0:
        return "SPARSE"
    return "NO RESULTS"


# ── EVENT INDEX sheet ──────────────────────────────────────────────────────────
# Shows ALL events (including sparse / no results).
# Year sheets show only pub_eids (non-sparse) events; EVENT INDEX is the complete record.

def build_event_index(wb: Workbook, events: dict, discs: dict,
                      raw_results, quarantine: set,
                      event_col_map: dict) -> None:
    ws = wb.create_sheet("EVENT INDEX")

    headers = ["Year", "Event Name", "Location", "Event Type",
               "Disciplines", "Placements", "Coverage", "Source", "Notes"]
    widths  = [6, 52, 30, 26, 12, 11, 12, 14, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "EVENT INDEX — ALL EVENTS", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    placements_per_event: dict[str, int] = Counter(
        r["event_key"] for r in raw_results
        if r.get("person_id", "").strip() not in _SKIP_PID
    )
    discs_per_event: dict[str, int] = Counter(k[0] for k in discs)

    sorted_events = sorted(events.values(),
                           key=lambda e: (e.get("year", ""),
                                          e.get("start_date", ""),
                                          e.get("event_name", "")))

    _COV_FILL = {
        "FULL":        None,
        "PARTIAL":     FILL_WARN,
        "SPARSE":      FILL_OTHER,
        "QUARANTINED": FILL_QUAR,
        "NO RESULTS":  FILL_OTHER,
    }

    shown = 0
    for ev in sorted_events:
        eid     = ev["event_key"]
        is_quar = eid in quarantine
        n_discs = discs_per_event.get(eid, 0)
        n_plc   = placements_per_event.get(eid, 0)
        src     = ev.get("source", "")
        cov     = _coverage_level(ev, n_discs, n_plc, is_quar)

        notes = "⛔ under review" if is_quar else ""

        if _is_worlds(ev):
            row_fill = FILL_WORLDS_ROW
        else:
            row_fill = _COV_FILL.get(cov)

        name_val  = _worlds_name_override(ev) or ev.get("event_name", eid)
        year_str  = ev.get("year", "")
        sheet_ref = event_col_map.get(eid)

        for col_idx, v in enumerate([
            year_str,
            name_val,
            _location(ev),
            ev.get("event_type", ""),
            n_discs or "",
            n_plc   or "",
            cov,
            src,
            notes,
        ], 1):
            font = FONT_DATA
            if col_idx == 2 and sheet_ref:
                font = FONT_LINK
            cell = _w(ws, row, col_idx, v, font=font, align=ALIGN_L)
            if row_fill:
                cell.fill = row_fill
            if col_idx == 2 and sheet_ref:
                yr_sheet, col_letter = sheet_ref
                safe = yr_sheet.replace("'", "''")
                try:
                    cell.hyperlink = f"#{safe}!{col_letter}1"
                    cell.style = "Hyperlink"
                except Exception:
                    cell.font = FONT_LINK

        row += 1
        shown += 1

    row += 1
    _w(ws, row, 1, "Coverage levels:", font=FONT_HEADER, align=ALIGN_L)
    _w(ws, row, 2,
       "SPARSE and NO RESULTS events are excluded from year sheets but listed here "
       "and in QC - EXCLUDED EVENTS.",
       font=FONT_NOTE, align=ALIGN_L)
    row += 1
    for level, desc, fill_key in [
        ("FULL",        "≥20 placements and ≥3 divisions",          "FULL"),
        ("PARTIAL",     "Some results but not fully complete",       "PARTIAL"),
        ("SPARSE",      "Fewer than 10 placements or 2 disciplines", "SPARSE"),
        ("NO RESULTS",  "No placement data available",               "NO RESULTS"),
        ("QUARANTINED", "Data under review — results may be uncertain", "QUARANTINED"),
        ("🌐 WORLDS",   "World Footbag Championship event",          "_WORLDS"),
    ]:
        f = _COV_FILL.get(fill_key) or (FILL_WORLDS_ROW if fill_key == "_WORLDS" else FILL_NONE)
        _w(ws, row, 1, f"  {level}", font=FONT_DATA, fill=f, align=ALIGN_L)
        _w(ws, row, 2, desc, font=FONT_DATA, align=ALIGN_L)
        row += 1

    print(f"  EVENT INDEX: {shown} events (all)")


# ── QC - EXCLUDED EVENTS sheet ────────────────────────────────────────────────

def build_excluded_events(wb: Workbook, events: dict, quarantine: set,
                          discs: dict, raw_results) -> None:
    ws = wb.create_sheet("QC - EXCLUDED EVENTS")

    headers = ["Year", "Event Key", "Event Name", "Location",
               "Coverage", "Discs", "Placements", "Reason"]
    widths  = [6, 28, 52, 30, 13, 7, 11, 36]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "QC — EVENTS EXCLUDED FROM YEAR SHEETS", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    placements_per_event: dict[str, int] = Counter(
        r["event_key"] for r in raw_results
        if r.get("person_id", "").strip() not in _SKIP_PID
    )
    discs_per_event: dict[str, int] = Counter(k[0] for k in discs)

    _COV_FILL = {
        "SPARSE":      FILL_OTHER,
        "QUARANTINED": FILL_QUAR,
        "NO RESULTS":  FILL_OTHER,
    }

    excluded = []
    for eid, ev in events.items():
        is_quar = eid in quarantine
        n_discs = discs_per_event.get(eid, 0)
        n_plc   = placements_per_event.get(eid, 0)
        cov     = _coverage_level(ev, n_discs, n_plc, is_quar)

        if cov == "NO RESULTS":
            reason = "no results in database"
        elif cov == "SPARSE":
            reason = f"sparse coverage ({n_plc} placements, {n_discs} discipline(s))"
        elif is_quar:
            reason = "under review (complex or uncertain data)"
        else:
            continue

        excluded.append((ev.get("year", ""), eid, ev.get("event_name", ""),
                         _location(ev), cov, n_discs or "", n_plc or "", reason))

    excluded.sort(key=lambda x: (x[0], x[2].lower()))
    for rec in excluded:
        cov_val  = rec[4]
        row_fill = _COV_FILL.get(cov_val)
        for col, v in enumerate(rec, 1):
            cell = _w(ws, row, col, v, font=FONT_DATA, align=ALIGN_L)
            if row_fill:
                cell.fill = row_fill
        row += 1

    print(f"  QC - EXCLUDED EVENTS: {len(excluded)} entries")


# ── Year sheets ────────────────────────────────────────────────────────────────

def _team_display(slot_rows: list[dict]) -> tuple[str, bool]:
    by_order: dict[int, str] = {}
    for r in slot_rows:
        try:
            order = int(r.get("participant_order", "1") or 1)
        except ValueError:
            order = 1
        name = r.get("display_name", "").strip()
        if name and name not in _SKIP_DNAME:
            by_order[order] = name
        elif not by_order.get(order):
            by_order[order] = _UNKNOWN_DISPLAY

    if not by_order:
        return (_UNKNOWN_DISPLAY, True)

    # Include ALL participants at this placement (handles ties with 3+ players)
    parts = [by_order[k] for k in sorted(by_order.keys())]

    parts = [p for p in parts if p]
    disp  = " / ".join(parts) if parts else _UNKNOWN_DISPLAY
    is_unk = all(p == _UNKNOWN_DISPLAY for p in parts)
    return (disp, is_unk)


def build_year_sheets(wb: Workbook, raw_results, events: dict,
                      discs: dict, persons: dict,
                      quarantine: set, pub_eids: set) -> dict:
    """Build one sheet per year. Returns event_col_map = {event_key: (sheet_name, col_letter)}.

    Only pub_eids (non-sparse) events appear in year sheets.
    Placements are keyed by discipline_name so all disc_keys sharing the same
    canonical name are merged into one row per event.
    Participants are deduplicated by person_id.
    """
    # plcmt_by_event[eid][disc_name][place] = merged list of participant rows
    plcmt_by_event: dict[str, dict] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    # div_can_rec[(eid, disc_name)] = first disc record seen (for sort key + fill colour)
    div_can_rec: dict[tuple, dict] = {}
    _raw_discs_per_can: dict[tuple, set] = defaultdict(set)

    for r in raw_results:
        eid      = r["event_key"]
        disc_key = r["discipline_key"]
        rec      = discs.get((eid, disc_key), {})
        disc_name = rec.get("discipline_name") or disc_key
        try:
            place = int(r.get("placement", "0") or 0)
        except ValueError:
            place = 0
        plcmt_by_event[eid][disc_name][place].append(r)
        _raw_discs_per_can[(eid, disc_name)].add(disc_key)
        if (eid, disc_name) not in div_can_rec:
            div_can_rec[(eid, disc_name)] = rec

    _dup_groups = sum(1 for v in _raw_discs_per_can.values() if len(v) > 1)
    _rows_merged = sum(len(v) - 1 for v in _raw_discs_per_can.values() if len(v) > 1)
    if _dup_groups:
        print(f"  Canonical merge: {_dup_groups} duplicate groups, "
              f"{_rows_merged} raw discipline rows consolidated")

    year_to_eids: dict[str, list] = defaultdict(list)
    for eid, ev in events.items():
        yr = ev.get("year", "").strip()
        if yr and yr.isdigit() and eid in pub_eids:
            year_to_eids[yr].append(eid)

    event_col_map: dict[str, tuple] = {}

    _R_NAME = 1
    _R_LOC  = 2
    _R_DATE = 3
    _R_TYPE = 4
    _R_EID  = 5
    _R_DATA = 7

    _YF_META = Font(bold=True, size=11)
    _YF_EID  = Font(size=8, color="808080")
    _YF_DIV  = Font(bold=True, size=10)
    _YF_PLC  = Font(size=10)
    _YF_UNK  = Font(size=10, italic=True, color="A0A0A0")
    _YF_QUAR = Font(bold=True, size=11, color="CC0000")

    _COL_A_W = 14
    _COL_MIN = 22
    _COL_MAX = 52
    _TOP_PLCS = 10

    for yr in sorted(year_to_eids, key=int):
        eids = sorted(
            year_to_eids[yr],
            key=lambda eid: (events[eid].get("start_date", "") or "",
                             events[eid].get("event_name", ""))
        )
        ws = wb.create_sheet(title=yr)
        ws.column_dimensions["A"].width = _COL_A_W

        label_map = {
            _R_NAME: "Event",
            _R_LOC:  "Location",
            _R_DATE: "Date",
            _R_TYPE: "Type",
            _R_EID:  "Event Key",
        }
        for r_idx, label in label_map.items():
            _w(ws, r_idx, 1, label,
               font=Font(size=9, color="606060"),
               align=Alignment(horizontal="right", vertical="top"))

        div_sort_keys_yr: dict[str, tuple] = {}
        for eid in eids:
            for disc_name in plcmt_by_event.get(eid, {}):
                if disc_name not in div_sort_keys_yr:
                    rec = div_can_rec.get((eid, disc_name), {})
                    div_sort_keys_yr[disc_name] = _disc_sort_key(disc_name, rec)

        all_discs_this_year: list[str] = sorted(
            div_sort_keys_yr, key=lambda d: div_sort_keys_yr[d]
        )

        disc_row_start: dict[str, int] = {}
        cur_row = _R_DATA
        for disc_name in all_discs_this_year:
            disc_row_start[disc_name] = cur_row
            cur_row += 1 + _TOP_PLCS

        for disc_name, dr in disc_row_start.items():
            _w(ws, dr, 1, disc_name[:22],
               font=Font(size=8, italic=True, color="808080"),
               align=Alignment(horizontal="right", vertical="top"))
            for pi in range(1, _TOP_PLCS + 1):
                _w(ws, dr + pi, 1, f"  p{pi}",
                   font=Font(size=8, color="A0A0A0"),
                   align=Alignment(horizontal="right", vertical="top"))

        for col_offset, eid in enumerate(eids, start=2):
            ev         = events[eid]
            is_quar    = eid in quarantine
            placements = plcmt_by_event.get(eid, {})

            col_letter = get_column_letter(col_offset)
            event_col_map[eid] = (yr, col_letter)

            ev_name = _worlds_name_override(ev) or ev.get("event_name", eid)
            if is_quar:
                ev_name = "⛔ " + ev_name

            _w(ws, _R_NAME, col_offset, ev_name,
               font=(_YF_QUAR if is_quar else _YF_META), align=ALIGN_LW)
            _w(ws, _R_LOC,  col_offset, _location(ev),           font=_YF_PLC, align=ALIGN_L)
            _w(ws, _R_DATE, col_offset, ev.get("start_date", ""), font=_YF_PLC, align=ALIGN_L)
            _w(ws, _R_TYPE, col_offset, ev.get("event_type", ""), font=_YF_PLC, align=ALIGN_L)
            _w(ws, _R_EID,  col_offset, eid,                      font=_YF_EID, align=ALIGN_L)

            anchor_name = f"event_{eid}"
            safe_yr = yr.replace("'", "''")
            try:
                dn = DefinedName(
                    name=anchor_name,
                    attr_text=f"'{safe_yr}'!${col_letter}$1",
                )
                wb.defined_names[anchor_name] = dn
            except Exception:
                pass

            max_name_len = len(ev_name)

            for disc_name, disc_placements in placements.items():
                if disc_name not in disc_row_start:
                    continue
                dr  = disc_row_start[disc_name]
                rec = div_can_rec.get((eid, disc_name), {})
                cat = rec.get("discipline_category", "")
                dfill = _cat_fill(cat)

                _w(ws, dr, col_offset, disc_name,
                   font=_YF_DIV, fill=dfill, align=ALIGN_L)
                max_name_len = max(max_name_len, len(disc_name))

                # Dense placement rendering: iterate only the actual place values
                # present in the data, writing to consecutive rows.  Tie-adjusted
                # rankings (e.g. 1,2,3,5,9) are packed without blank holes.
                # When a gap exists the actual place number is embedded in the cell
                # text (e.g. "5. Aleksi Airinen / ...") so the reader knows the
                # official finishing position.
                actual_places = sorted(disc_placements.keys())
                for disp_idx, pi in enumerate(actual_places[:_TOP_PLCS], start=1):
                    slot_rows = _dedup_slot(disc_placements[pi])
                    disp, is_unk = _team_display(slot_rows)
                    if disp and pi != disp_idx:
                        disp = f"{pi}. {disp}"
                    font = _YF_UNK if is_unk and disp else _YF_PLC
                    _w(ws, dr + disp_idx, col_offset, disp, font=font, align=ALIGN_L)
                    max_name_len = max(max_name_len, len(disp))

            ws.column_dimensions[col_letter].width = min(
                max(_COL_MIN, max_name_len + 2), _COL_MAX
            )

        ws.row_dimensions[_R_NAME].height = 32
        ws.row_dimensions[_R_LOC].height  = 14
        ws.row_dimensions[_R_DATE].height = 14
        ws.row_dimensions[_R_TYPE].height = 14
        ws.row_dimensions[_R_EID].height  = 11
        ws.freeze_panes = "B6"

        n_q = sum(1 for e in eids if e in quarantine)
        print(f"  {yr}: {len(eids)} events"
              + (f" ({n_q} quarantined)" if n_q else ""))

    return event_col_map


# ── Validation ─────────────────────────────────────────────────────────────────

def validate_workbook(wb: Workbook) -> dict:
    issues = []
    hidden_sheets = hidden_cols = hidden_rows = 0

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            hidden_sheets += 1
            issues.append(f"HIDDEN SHEET: {ws.title}")

        cd = ws.column_dimensions.get("A")
        if cd and (cd.hidden or (cd.width is not None and cd.width <= 0)):
            hidden_cols += 1
            issues.append(f"HIDDEN/ZERO COL A: {ws.title}")

        for cd in ws.column_dimensions.values():
            if cd.hidden:
                hidden_cols += 1
                issues.append(f"HIDDEN COL: {ws.title}")

        for rd in ws.row_dimensions.values():
            if rd.hidden:
                hidden_rows += 1
                issues.append(f"HIDDEN ROW {rd.index}: {ws.title}")

    return {
        "hidden_sheets": hidden_sheets,
        "hidden_cols":   hidden_cols,
        "hidden_rows":   hidden_rows,
        "issues":        issues,
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"\nbuild_workbook_release.py  ({VERSION} · {UPDATED})")
    print(f"  Source:  {CANONICAL_INPUT}")
    print(f"  Output:  {OUTPUT_PATH}\n")

    events, discs, raw_results, persons, quarantine = load_all()

    # Compute publication set: non-sparse events go into year sheets.
    # EVENT INDEX always shows all events.
    pub_eids = compute_pub_eids(events, discs, raw_results, quarantine)
    print(f"  Publication set (year sheets): {len(pub_eids)} of {len(events)} events")

    print("Computing statistics…")
    stats = compute_stats(raw_results, events, discs, persons)
    print(f"  Persons with stats: {len(stats)}")

    wb = Workbook()
    wb.remove(wb.active)

    print("\nBuilding sheets…")

    build_readme(wb, events, persons, len(raw_results))
    build_statistics(wb, stats, persons)
    build_era_leaders(wb, raw_results, events, discs, persons)
    build_player_stats(wb, stats, persons)

    event_col_map = build_year_sheets(
        wb, raw_results, events, discs, persons, quarantine, pub_eids)

    build_event_index(wb, events, discs, raw_results, quarantine, event_col_map)
    build_excluded_events(wb, events, quarantine, discs, raw_results)

    # Reorder front-matter sheets before year sheets
    desired_front = ["README", "EVENT INDEX", "STATISTICS", "ERA LEADERS",
                     "PLAYER STATS", "QC - EXCLUDED EVENTS"]
    final_order = desired_front + [s for s in wb.sheetnames if s not in desired_front]
    for i, name in enumerate(final_order):
        if name in wb.sheetnames:
            current = wb.sheetnames.index(name)
            if current != i:
                wb.move_sheet(name, offset=i - current)

    print("\nValidating…")
    v = validate_workbook(wb)
    print(f"  Hidden sheets:  {v['hidden_sheets']}")
    print(f"  Hidden columns: {v['hidden_cols']}")
    print(f"  Hidden rows:    {v['hidden_rows']}")
    if v["issues"]:
        for iss in v["issues"][:10]:
            print(f"  ⚠  {iss}")
    else:
        print("  ✓ No hidden structure found")

    print(f"\nSaving → {OUTPUT_PATH.name}…")
    wb.save(str(OUTPUT_PATH))
    print(f"  Done.  ({OUTPUT_PATH.stat().st_size / 1_048_576:.1f} MB)")

    print("\n── Summary ──────────────────────────────────────────────")
    year_sheets = [s for s in wb.sheetnames if s.isdigit()]
    print(f"  Sheets total:       {len(wb.sheetnames)}")
    print(f"  Front-matter:       {len(desired_front)}")
    print(f"  Year sheets:        {len(year_sheets)}"
          + (f"  ({year_sheets[0]}–{year_sheets[-1]})" if year_sheets else ""))
    print(f"  Events in year sheets: {len(event_col_map)}")
    print(f"  Events in EVENT INDEX: {len(events)}")
    print(f"  Persons in stats:   {len(stats)}")
    print(f"  Persons in PLAYER STATS: {len(persons)}")
    print(f"  Hidden structure:   {v['hidden_sheets'] + v['hidden_cols'] + v['hidden_rows']}")
    print(f"  Version:            {VERSION}")
    print(f"  Updated:            {UPDATED}")


if __name__ == "__main__":
    main()
