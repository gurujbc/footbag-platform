#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(".")
XLSX = ROOT / "Footbag_Results_Canonicalv1.xlsx"   # change if needed
CSV  = ROOT / "out" / "location_canon_full_final.csv"

def fmt_from_csv_row(r: dict) -> str:
    city = (r.get("city_canon") or "").strip()
    state = (r.get("state_canon") or "").strip()
    country = (r.get("country_canon") or "").strip()
    iso3 = (r.get("country_iso3") or "").strip().upper()

    if not (city or state or country):
        return ""

    # Match 03_build_excel.py behavior: USA formatting is based on iso3 == "USA"
    if iso3 == "USA":
        parts = [p for p in [city, state, country] if p]
        return ", ".join(parts)

    # non-USA
    parts = [p for p in [city, country] if p]
    return ", ".join(parts)

def load_loc_map(csv_path: Path) -> dict[str, dict]:
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8").fillna("")
    need = ["event_id", "city_canon", "state_canon", "country_canon", "country_iso3"]
    missing = [c for c in need if c not in df.columns]
    if missing:
        raise SystemExit(f"CSV missing required columns {missing} in {csv_path}")

    m: dict[str, dict] = {}
    for r in df.to_dict(orient="records"):
        eid = str(r["event_id"]).strip()
        if not eid:
            continue
        m[eid] = r
    return m

def get_year_sheets(wb) -> list[str]:
    out = []
    for name in wb.sheetnames:
        if name.isdigit() and len(name) == 4:
            out.append(name)
    return sorted(out)

def read_location_from_year_sheet(ws) -> str:
    # Your year sheets use labels in col A, value in col B.
    # Row 4 should be: A4="Location", B4=<value>
    a4 = ws["A4"].value
    b4 = ws["B4"].value
    if isinstance(a4, str) and a4.strip().lower() == "location":
        return "" if b4 is None else str(b4).strip()
    # fallback: scan first ~15 rows for "Location" label
    for r in range(1, 16):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().lower() == "location":
            v = ws.cell(row=r, column=2).value
            return "" if v is None else str(v).strip()
    return ""

def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing XLSX: {XLSX}")
    if not CSV.exists():
        raise SystemExit(f"Missing CSV: {CSV} (this is what 03 expects)")

    loc_map = load_loc_map(CSV)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # --- Index sheet check (if it has event_id + Location)
    if "Index" in wb.sheetnames:
        ws = wb["Index"]
        # Find header row by scanning first 10 rows for "event_id"
        header_row = None
        header = {}
        for r in range(1, 11):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 40)]
            if any(isinstance(v, str) and v.strip().lower() == "event_id" for v in vals):
                header_row = r
                for c, v in enumerate(vals, start=1):
                    if isinstance(v, str) and v.strip():
                        header[v.strip()] = c
                break

        idx_mismatches = []
        if header_row and "event_id" in header and ("Location" in header or "location" in header):
            loc_col = header.get("Location") or header.get("location")
            eid_col = header["event_id"]
            for r in range(header_row + 1, header_row + 1 + 2000):  # plenty
                eid = ws.cell(row=r, column=eid_col).value
                if eid is None:
                    break
                eid = str(eid).strip()
                shown = ws.cell(row=r, column=loc_col).value
                shown = "" if shown is None else str(shown).strip()

                canon_row = loc_map.get(eid)
                if not canon_row:
                    continue
                expected = fmt_from_csv_row(canon_row).strip()
                if expected and shown and shown != expected:
                    idx_mismatches.append((eid, shown, expected))

        print("CSV rows:", len(loc_map))
        if header_row:
            print("Index header row:", header_row, "columns found:", sorted(header.keys())[:12], "...")
        if idx_mismatches:
            print("\nIndex mismatches (first 20):")
            for eid, shown, exp in idx_mismatches[:20]:
                print(" ", eid, "| shown:", shown, "| expected:", exp)
        else:
            print("\nIndex mismatches: 0 (or Index has no Location column / couldn’t detect)")

    # --- Year sheet check
    year_sheets = get_year_sheets(wb)
    mismatches = []
    missing_in_csv = 0
    blank_expected = 0

    for ys in year_sheets:
        ws = wb[ys]
        eid = ws["B1"].value  # row 1 event_id in col B
        if eid is None:
            continue
        eid = str(eid).strip()

        shown = read_location_from_year_sheet(ws)
        canon_row = loc_map.get(eid)
        if not canon_row:
            missing_in_csv += 1
            continue
        expected = fmt_from_csv_row(canon_row).strip()
        if not expected:
            blank_expected += 1
            continue
        if shown != expected:
            mismatches.append((ys, eid, shown, expected))

    print("\nYear sheets:", len(year_sheets))
    print("Year sheets missing from CSV:", missing_in_csv)
    print("Year sheets with blank expected canon (csv had no city/country):", blank_expected)
    print("Year-sheet mismatches:", len(mismatches))

    if mismatches:
        print("\nSample mismatches (first 30):")
        for ys, eid, shown, exp in mismatches[:30]:
            print(f" {ys} {eid} | shown: {shown} | expected: {exp}")

    # Extra diagnostic: iso3 coverage for USA formatting
    df = pd.read_csv(CSV, dtype=str, encoding="utf-8").fillna("")
    usa_like = df[df["country_canon"].str.strip().str.lower().eq("united states")]
    usa_iso3 = usa_like[usa_like["country_iso3"].str.strip().str.upper().eq("USA")]
    print("\nUSA rows where country_canon='United States':", len(usa_like))
    print("...of those, iso3=='USA':", len(usa_iso3))
    print("(If this is low, 03 will NOT apply City, State, Country formatting reliably.)")

if __name__ == "__main__":
    main()
