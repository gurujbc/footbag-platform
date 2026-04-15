#!/usr/bin/env python3
from pathlib import Path
import csv
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path.home() / "projects" / "FOOTBAG_DATA"
INPUT_XLSX = ROOT / "Footbag_Results_Merged_FINAL.xlsx"
SUMMARY_CSV = ROOT / "early_data" / "out" / "records" / "records_summary.csv"
LEADERBOARD_CSV = ROOT / "early_data" / "out" / "records" / "records_leaderboard_by_trick.csv"
OUTPUT_XLSX = ROOT / "Footbag_Results_Merged_WITH_RECORDS.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEAD_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FONT = Font(size=15, bold=True)
BOLD = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
THIN = Side(style="thin", color="B7C9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def style_header(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def style_subhead(cell):
    cell.fill = SUBHEAD_FILL
    cell.font = BOLD
    cell.border = BORDER


def set_widths(ws, widths):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_table(ws, start_row, start_col, end_row, end_col, name):
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def insert_sheet_after(wb, new_sheet_name, after_sheet_name):
    ws = wb.create_sheet(new_sheet_name)
    idx = wb.sheetnames.index(after_sheet_name)
    wb._sheets.remove(ws)
    wb._sheets.insert(idx + 1, ws)
    return ws


def write_hyperlink(cell, url, label="Video"):
    if url:
        cell.value = label
        cell.hyperlink = url
        cell.font = LINK_FONT


def build_records_notes(ws):
    ws["A1"] = "RECORDS NOTES"
    ws["A1"].font = TITLE_FONT

    notes = [
        "These records are a separate layer from event results.",
        "They are primarily trick-based freestyle records sourced from Passback and linked to the player identity system where possible.",
        "Records may be video-linked, community-sourced, or flagged for review.",
        "Confidence and verification status should be interpreted separately from canonical event placements.",
        "The existing CONSECUTIVE RECORDS sheet remains event-derived; TRICK RECORDS is cross-event and trick-specific.",
    ]

    row = 3
    for note in notes:
        ws[f"A{row}"] = f"- {note}"
        row += 1

    ws["A10"] = "Suggested interpretation:"
    ws["A10"].font = BOLD
    ws["A11"] = "Use TRICK RECORDS for per-trick leaderboards, RECORDS OVERVIEW for summary highlights, and year sheets / event index for canonical competition history."

    set_widths(ws, {1: 120})
    ws.freeze_panes = "A3"


def build_trick_records(ws, leaderboard_rows):
    ws["A1"] = "TRICK RECORDS"
    ws["A1"].font = TITLE_FONT

    headers = [
        "Trick", "Sort Name", "Adds", "Rank", "Player", "Player ID",
        "Record", "Date", "Video", "Time Clip", "Record ID"
    ]
    start_row = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, 1, len(headers))

    row = start_row + 1
    for r in leaderboard_rows:
        ws.cell(row=row, column=1, value=r["trick_name"])
        ws.cell(row=row, column=2, value=r["sort_name"])
        ws.cell(row=row, column=3, value=int(r["adds"]) if str(r["adds"]).isdigit() else r["adds"])
        ws.cell(row=row, column=4, value=int(r["rank"]) if str(r["rank"]).isdigit() else r["rank"])
        ws.cell(row=row, column=5, value=r["player"])
        ws.cell(row=row, column=6, value=r["player_id"])
        ws.cell(row=row, column=7, value=float(r["record_value"]) if r["record_value"] else None)
        ws.cell(row=row, column=8, value=r["date_normalized"])

        video_cell = ws.cell(row=row, column=9)
        write_hyperlink(video_cell, r["video"], "Video")

        ws.cell(row=row, column=10, value=r["time_clip"])

        record_id_cell = ws.cell(row=row, column=11)
        if r["video"]:
            write_hyperlink(record_id_cell, r["video"], r["record_id"])
        else:
            record_id_cell.value = r["record_id"]

        row += 1

    end_row = row - 1
    for r in range(start_row + 1, end_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    set_widths(ws, {
        1: 26, 2: 30, 3: 8, 4: 8, 5: 22, 6: 18,
        7: 10, 8: 12, 9: 12, 10: 10, 11: 30
    })
    ws.freeze_panes = "A4"
    add_table(ws, start_row, 1, end_row, len(headers), "TrickRecordsTable")


def build_records_overview(ws, summary_rows, leaderboard_rows):
    ws["A1"] = "RECORDS OVERVIEW"
    ws["A1"].font = TITLE_FONT

    total_records = len(leaderboard_rows)
    unique_tricks = len(summary_rows)
    players = [r["player"] for r in leaderboard_rows if r["player"]]
    unique_players = len(set(players))
    player_counts = Counter(players)

    ws["A3"] = "Key Metrics"
    style_subhead(ws["A3"])
    ws["A4"] = "Total Record Rows"
    ws["B4"] = total_records
    ws["A5"] = "Unique Tricks"
    ws["B5"] = unique_tricks
    ws["A6"] = "Players Represented"
    ws["B6"] = unique_players

    for r in range(4, 7):
        ws[f"A{r}"].font = BOLD
        ws[f"A{r}"].border = BORDER
        ws[f"B{r}"].border = BORDER

    ws["D3"] = "Top Players by Record Entries"
    style_subhead(ws["D3"])
    ws["D4"] = "Player"
    ws["E4"] = "# Records"
    style_header(ws, 4, 4, 5)

    row = 5
    for player, cnt in player_counts.most_common(10):
        ws.cell(row=row, column=4, value=player)
        ws.cell(row=row, column=5, value=cnt)
        ws.cell(row=row, column=4).border = BORDER
        ws.cell(row=row, column=5).border = BORDER
        row += 1

    ws["A9"] = "Selected Notable Records"
    style_subhead(ws["A9"])
    note_headers = ["Trick", "Record", "Holder", "Date", "Video"]
    for i, h in enumerate(note_headers, start=1):
        ws.cell(row=10, column=i, value=h)
    style_header(ws, 10, 1, 5)

    ranked_summary = sorted(
        [r for r in summary_rows if r["top_record_value"]],
        key=lambda x: float(x["top_record_value"]),
        reverse=True
    )[:12]

    leaderboard_by_record_id = {r["record_id"]: r for r in leaderboard_rows}

    row = 11
    for r in ranked_summary:
        ws.cell(row=row, column=1, value=r["trick_name"])
        ws.cell(row=row, column=2, value=float(r["top_record_value"]))
        ws.cell(row=row, column=3, value=r["top_holder"])
        ws.cell(row=row, column=4, value=r["latest_date"] or r["first_date"])

        video_cell = ws.cell(row=row, column=5)
        top_id = r.get("top_record_id", "")
        video_url = ""
        if top_id and top_id in leaderboard_by_record_id:
            video_url = leaderboard_by_record_id[top_id].get("video", "")
        write_hyperlink(video_cell, video_url, "Video")

        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BORDER
        row += 1

    ws["D16"] = "Notes"
    style_subhead(ws["D16"])
    notes = [
        "Records are separate from canonical event results.",
        "These sheets summarize trick-based records sourced from Passback.",
        "Player IDs are linked where deterministically matched.",
        "Use TRICK RECORDS for full ranked per-trick detail.",
    ]
    r = 17
    for n in notes:
        ws.cell(row=r, column=4, value=f"- {n}")
        r += 1

    set_widths(ws, {1: 26, 2: 12, 3: 22, 4: 18, 5: 12})
    ws.freeze_panes = "A4"


def main():
    wb = load_workbook(INPUT_XLSX)

    for name in ["RECORDS OVERVIEW", "TRICK RECORDS", "RECORDS NOTES"]:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    summary_rows = read_csv(SUMMARY_CSV)
    leaderboard_rows = read_csv(LEADERBOARD_CSV)

    ws_notes = insert_sheet_after(wb, "RECORDS NOTES", "STATISTICS")
    ws_trick = insert_sheet_after(wb, "TRICK RECORDS", "STATISTICS")
    ws_overview = insert_sheet_after(wb, "RECORDS OVERVIEW", "STATISTICS")

    build_records_overview(ws_overview, summary_rows, leaderboard_rows)
    build_trick_records(ws_trick, leaderboard_rows)
    build_records_notes(ws_notes)

    wb.save(OUTPUT_XLSX)
    print(f"Saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
